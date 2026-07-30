import hashlib
import io
import json
import re
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo


st.set_page_config(
    page_title="Manpower AI Agent",
    page_icon="🏗️",
    layout="wide",
)

REQUIRED_TRADES = ("AC", "EL", "FS", "PD")
APP_VERSION = "0.7.0 — New project Excel"
PARSER_MODE_OPTIONS = {
    "自動偵測": "auto",
    "Location＋Manpower數字欄": "numeric_table",
    "工人姓名表（每列1人）": "worker_table",
    "工作描述中的X人／姓名數量": "description",
    "只使用報告總人數": "total_only",
}


# =========================================================
# General helper functions
# =========================================================


def text_to_list(value: str) -> list[str]:
    return [
        item.strip()
        for item in value.replace("，", ",").split(",")
        if item.strip()
    ]


def normalize_floor(value: str) -> str:
    original = value.strip().upper()
    compact = re.sub(r"\s+", "", original)

    if re.fullmatch(r"G/?F", compact):
        return "GF"

    basement_match = re.fullmatch(r"B(\d+)/?F?", compact)
    if basement_match:
        return f"B{int(basement_match.group(1))}"

    floor_match = re.fullmatch(r"(\d+)/?F?", compact)
    if floor_match:
        return f"{int(floor_match.group(1))}F"

    return original


def floor_sort_key(value: str) -> tuple[int, str]:
    match = re.fullmatch(r"(\d+)F", value)
    if match:
        return int(match.group(1)), value
    return 9999, value


def calculate_file_hash(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def detect_trade_from_filename(filename: str) -> str | None:
    filename_upper = filename.upper()
    trade_patterns = {
        "AC": [r"(^|[^A-Z0-9])AC([^A-Z0-9]|$)", r"MVAC"],
        "EL": [r"(^|[^A-Z0-9])EL([^A-Z0-9]|$)", r"ELECTRICAL"],
        "FS": [r"(^|[^A-Z0-9])FS([^A-Z0-9]|$)", r"FIRE[\s_-]*SERVICE"],
        "PD": [r"(^|[^A-Z0-9])PD([^A-Z0-9]|$)", r"PLUMBING", r"DRAINAGE"],
    }

    for trade, patterns in trade_patterns.items():
        if any(re.search(pattern, filename_upper) for pattern in patterns):
            return trade

    return None


def format_file_size(size_bytes: int) -> str:
    if size_bytes < 1024:
        return f"{size_bytes} B"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes / (1024 * 1024):.1f} MB"


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def location_sort_key(location: str) -> tuple[int, int, int, str]:
    """Sort locations in site order rather than alphabetically.

    Order:
    1. T1, T2, T3... with floors from highest to lowest
    2. Podium with numbered floors from highest to lowest, then GF
    3. Basement from B1 downward to B2, B3, B4...
    4. Special and uncertain locations
    """
    text = clean_cell(location).upper()

    tower_floor_match = re.fullmatch(
        r"T\s*(\d+)\s*/\s*(\d+)F",
        text,
    )
    if tower_floor_match:
        tower_number = int(tower_floor_match.group(1))
        floor_number = int(tower_floor_match.group(2))
        return (0, tower_number, -floor_number, text)

    podium_floor_match = re.fullmatch(
        r"PODIUM\s*/\s*(GF|\d+F)",
        text,
    )
    if podium_floor_match:
        floor = podium_floor_match.group(1)
        floor_rank = (
            0
            if floor == "GF"
            else -int(re.search(r"\d+", floor).group())
        )
        return (1, 0, floor_rank, text)

    basement_floor_match = re.fullmatch(
        r"BASEMENT\s*/\s*B(\d+)",
        text,
    )
    if basement_floor_match:
        basement_number = int(basement_floor_match.group(1))
        return (2, 0, basement_number, text)

    tower_unspecified_match = re.fullmatch(
        r"T\s*(\d+)\s*/\s*FLOOR\s*U",
        text,
    )
    if tower_unspecified_match:
        return (
            4,
            int(tower_unspecified_match.group(1)),
            0,
            text,
        )

    if text == "PODIUM / FLOOR U":
        return (5, 0, 0, text)

    if text == "BASEMENT / FLOOR U":
        return (6, 0, 0, text)

    cross_tower_match = re.match(
        r"CROSS-FLOOR\s*/\s*T\s*(\d+)",
        text,
    )
    if cross_tower_match:
        return (
            7,
            int(cross_tower_match.group(1)),
            0,
            text,
        )

    if text.startswith("CROSS-FLOOR"):
        return (8, 0, 0, text)

    if text.startswith("DISTRIBUTION U"):
        return (9, 0, 0, text)

    if text.startswith("UNSPECIFIED TOWER"):
        return (10, 0, 0, text)

    if text == "UNSPECIFIED":
        return (11, 0, 0, text)

    # Roof and other configured special locations come after the main zones.
    return (3, 0, 0, text)


def build_location_summary(
    detail_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Merge manpower by date, standard location and trade."""
    empty_columns = [
        "日期",
        "標準位置",
        "AC",
        "EL",
        "FS",
        "PD",
        "合計",
    ]

    if detail_df.empty:
        empty_df = pd.DataFrame(columns=empty_columns)
        return empty_df, empty_df.copy()

    working_df = detail_df.copy()

    required_columns = {
        "日期": "",
        "工種": "",
        "標準位置": "Unspecified",
        "人數": 0,
        "位置狀態": "需人工確認位置",
    }

    for column_name, default_value in required_columns.items():
        if column_name not in working_df.columns:
            working_df[column_name] = default_value

    working_df["日期"] = (
        working_df["日期"]
        .fillna("")
        .astype(str)
        .str.strip()
    )
    working_df["工種"] = (
        working_df["工種"]
        .fillna("")
        .astype(str)
        .str.upper()
        .str.strip()
    )
    working_df["標準位置"] = (
        working_df["標準位置"]
        .fillna("Unspecified")
        .astype(str)
        .str.strip()
        .replace("", "Unspecified")
    )
    working_df["位置狀態"] = (
        working_df["位置狀態"]
        .fillna("需人工確認位置")
        .astype(str)
        .str.strip()
    )
    working_df["人數"] = pd.to_numeric(
        working_df["人數"],
        errors="coerce",
    ).fillna(0)

    working_df = working_df[
        working_df["人數"] > 0
    ].copy()

    def create_summary(source_df: pd.DataFrame) -> pd.DataFrame:
        if source_df.empty:
            return pd.DataFrame(columns=empty_columns)

        grouped_df = (
            source_df.groupby(
                ["日期", "標準位置", "工種"],
                dropna=False,
                as_index=False,
            )["人數"]
            .sum()
        )

        pivot_df = grouped_df.pivot_table(
            index=["日期", "標準位置"],
            columns="工種",
            values="人數",
            aggfunc="sum",
            fill_value=0,
        ).reset_index()

        pivot_df.columns.name = None

        for trade in REQUIRED_TRADES:
            if trade not in pivot_df.columns:
                pivot_df[trade] = 0

        for trade in REQUIRED_TRADES:
            pivot_df[trade] = (
                pd.to_numeric(
                    pivot_df[trade],
                    errors="coerce",
                )
                .fillna(0)
                .astype(int)
            )

        pivot_df["合計"] = pivot_df[
            list(REQUIRED_TRADES)
        ].sum(axis=1)

        pivot_df = pivot_df[
            [
                "日期",
                "標準位置",
                *REQUIRED_TRADES,
                "合計",
            ]
        ]

        ordered_index = sorted(
            pivot_df.index,
            key=lambda index: (
                str(pivot_df.at[index, "日期"]),
                location_sort_key(
                    str(pivot_df.at[index, "標準位置"])
                ),
            ),
        )

        return pivot_df.loc[
            ordered_index
        ].reset_index(drop=True)

    confirmed_mask = (
        working_df["位置狀態"] == "已解析"
    )

    confirmed_summary = create_summary(
        working_df[confirmed_mask]
    )
    review_summary = create_summary(
        working_df[~confirmed_mask]
    )

    return confirmed_summary, review_summary


# =========================================================
# PDF and EL parsing functions
# =========================================================


def extract_pdf_text(file_bytes: bytes) -> str:
    page_texts: list[str] = []

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
            page_texts.append(f"--- Page {page_number} ---\n{text}")

    return "\n\n".join(page_texts)


def find_first(pattern: str, text: str, flags: int = 0) -> str | None:
    match = re.search(pattern, text, flags)
    return match.group(1).strip() if match else None


def extract_header_data(text: str) -> dict[str, Any]:
    date = find_first(r"Date:\s*(\d{4}-\d{2}-\d{2})", text, re.I)
    trade = find_first(r"Trade:\s*([^\n]+?)(?:\s+Ref\.|\n)", text, re.I)
    reference = find_first(r"Ref\.\s*No\.:?\s*([^\s\n]+)", text, re.I)
    worker_text = find_first(r"\bWorker\s+(\d+)\b", text, re.I)
    total_text = find_first(
        r"Today\s+Total\s+Manpower\s*\*?\s*(\d+)\b",
        text,
        re.I,
    )

    worker = int(worker_text) if worker_text is not None else None
    total = int(total_text) if total_text is not None else None

    return {
        "date": date,
        "trade": trade,
        "reference": reference,
        "worker": worker,
        "today_total": total,
        "management_staff": (
            total - worker
            if total is not None and worker is not None
            else None
        ),
    }


def extract_site_work_table_rows(file_bytes: bytes) -> list[dict[str, str]]:
    extracted_rows: list[dict[str, str]] = []

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []

            for table in tables:
                if not table:
                    continue

                cleaned_table = [
                    [clean_cell(cell) for cell in row]
                    for row in table
                    if row
                ]

                header_index = None
                description_index = None
                location_index = None
                item_index = None

                for row_index, row in enumerate(cleaned_table):
                    upper_row = [cell.upper() for cell in row]

                    description_candidates = [
                        index
                        for index, cell in enumerate(upper_row)
                        if "DESCRIPTION" in cell
                    ]
                    location_candidates = [
                        index
                        for index, cell in enumerate(upper_row)
                        if "LOCATION" in cell
                    ]

                    if description_candidates and location_candidates:
                        header_index = row_index
                        description_index = description_candidates[0]
                        location_index = location_candidates[0]
                        item_candidates = [
                            index
                            for index, cell in enumerate(upper_row)
                            if cell == "ITEM"
                        ]
                        item_index = item_candidates[0] if item_candidates else 0
                        break

                if header_index is None:
                    continue

                for row in cleaned_table[header_index + 1 :]:
                    required_length = max(
                        item_index or 0,
                        description_index or 0,
                        location_index or 0,
                    ) + 1

                    if len(row) < required_length:
                        row += [""] * (required_length - len(row))

                    item = row[item_index or 0]
                    description = row[description_index or 0]
                    location = row[location_index or 0]

                    if not description and not location:
                        continue

                    if item and not re.fullmatch(r"1\.\d+", item):
                        continue

                    extracted_rows.append(
                        {
                            "item": item,
                            "description": description,
                            "location": location,
                        }
                    )

    return extracted_rows


def extract_fallback_detail_rows(text: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen_lines: set[str] = set()

    for line in text.splitlines():
        cleaned = clean_cell(line)

        if not cleaned or cleaned in seen_lines:
            continue

        has_explicit_people = bool(re.search(r"\d+\s*人", cleaned))
        has_name_list = bool(
            re.search(r"[：:]\s*[\u3400-\u9fff]{2,4}(?:\s*[,，、]\s*[\u3400-\u9fff]{2,4})+", cleaned)
        )

        if not has_explicit_people and not has_name_list:
            continue

        if "MANPOWER" in cleaned.upper() or "TOTAL" in cleaned.upper():
            continue

        seen_lines.add(cleaned)
        rows.append(
            {
                "item": "",
                "description": cleaned,
                "location": "",
            }
        )

    return rows


def parse_description_manpower(description: str) -> tuple[int | None, str]:
    explicit_counts = [
        int(value)
        for value in re.findall(r"(\d+)\s*人", description)
    ]

    if explicit_counts:
        return sum(explicit_counts), "報告列明人數"

    parts = re.split(r"[：:]", description, maxsplit=1)
    if len(parts) != 2:
        return None, "無法計算"

    possible_names = re.split(r"[,，、]", parts[1])
    normalized_names = [
        re.sub(r"\s+", "", name.strip())
        for name in possible_names
    ]
    valid_names = [
        name
        for name in normalized_names
        if re.fullmatch(r"[\u3400-\u9fff]{2,4}", name)
    ]

    if valid_names:
        return len(valid_names), "按姓名數量計算"

    return None, "無法計算"


def extract_towers(text: str) -> list[str]:
    normalized = text.upper().replace("Ｔ", "T")
    tower_numbers: set[int] = set()

    for group in re.findall(
        r"((?:\d+\s*[,，/&、]\s*)+\d+)\s*座",
        normalized,
    ):
        for number in re.findall(r"\d+", group):
            tower_numbers.add(int(number))

    for group in re.findall(
        r"T\s*(\d+(?:\s*[,/&、]\s*\d+)+)",
        normalized,
    ):
        for number in re.findall(r"\d+", group):
            tower_numbers.add(int(number))

    for number in re.findall(r"(?<!\d)(\d+)\s*座", normalized):
        tower_numbers.add(int(number))

    for number in re.findall(r"\bT\s*(\d+)\b", normalized):
        tower_numbers.add(int(number))

    return [f"T{number}" for number in sorted(tower_numbers)]


def normalize_report_location(value: str) -> str:
    """Normalise floor notation without collapsing multi-floor references."""
    text = clean_cell(value).upper()

    if not text:
        return ""

    text = (
        text.replace("地下", "GF")
        .replace("樓", "F")
        .replace("層", "F")
        .replace("－", "-")
        .replace("–", "-")
        .replace("—", "-")
        .replace("，", ",")
        .replace("、", ",")
        .replace("＆", "&")
    )

    # Convert G/F even when it appears inside a list such as G/F, 1F.
    text = re.sub(r"(?<![A-Z0-9])G\s*/\s*F(?![A-Z0-9])", "GF", text)

    # Convert B4/F and 6/F forms.
    text = re.sub(r"B\s*(\d+)\s*/\s*F", r"B\1", text)
    text = re.sub(r"(?<![A-Z])(?<!\d)(\d+)\s*/\s*F", r"\1F", text)

    text = re.sub(r"\s+", "", text)

    if text in {"N/A", "NA", "NIL", "NONE"}:
        return "NA"

    basement = re.fullmatch(r"B(\d+)F?", text)
    if basement:
        return f"B{int(basement.group(1))}"

    numbered_floor = re.fullmatch(r"(\d+)F", text)
    if numbered_floor:
        return f"{int(numbered_floor.group(1))}F"

    if text in {"G", "G/F", "GF"}:
        return "GF"

    return text


def extract_floor_tokens(text: str) -> list[str]:
    """Extract every floor token, including compressed lists such as 5,8F."""
    normalized = normalize_report_location(text)

    if normalized in {"", "NA"}:
        return []

    tokens: list[str] = []

    def add_floor(floor: str) -> None:
        if floor not in tokens:
            tokens.append(floor)

    if re.search(r"(?<![A-Z0-9])GF(?![A-Z0-9])", normalized):
        add_floor("GF")

    for number in re.findall(r"B(\d+)", normalized):
        add_floor(f"B{int(number)}")

    # Compressed forms where the final F applies to the whole list:
    # 5,8F / 2&3F / 2/3F. Parse these first to preserve source order.
    compressed_groups = re.findall(
        r"(?<![A-Z0-9])((?:\d+[,/&+]\s*)+\d+)F(?![A-Z0-9])",
        normalized,
    )
    for group in compressed_groups:
        for number in re.findall(r"\d+", group):
            add_floor(f"{int(number)}F")

    # Explicit forms: 2F, 6F, 25F.
    for number in re.findall(
        r"(?<![A-Z0-9])(\d+)F(?![A-Z0-9])",
        normalized,
    ):
        add_floor(f"{int(number)}F")

    return tokens


def build_location_result(
    description: str,
    location_raw: str,
    config: dict[str, Any],
) -> tuple[str, str, str, str]:
    canonical_location = normalize_report_location(location_raw)
    combined_text = f"{description} {location_raw}"
    towers = extract_towers(combined_text)
    floors = extract_floor_tokens(location_raw)

    if not floors:
        floors = extract_floor_tokens(description)

    tower_text = ", ".join(towers) if towers else ""
    floor_text = ", ".join(floors) if floors else ""

    if canonical_location == "NA" or not floors:
        return "Unspecified", tower_text, floor_text, "需人工確認位置"

    is_cross_floor = (
        len(floors) > 1
        or any(separator in canonical_location for separator in ["-", "+", "至", "&"])
    )

    if is_cross_floor:
        # Preserve compact source ranges. Example: the EL report location
        # "B4-3" means a grouped B4-to-3F record; it must not collapse to B4.
        basement_to_floor = re.fullmatch(
            r"B(\d+)-(\d+)F?",
            canonical_location,
            re.I,
        )
        basement_range = re.fullmatch(
            r"B(\d+)-B(\d+)",
            canonical_location,
            re.I,
        )

        if basement_to_floor:
            readable_floors = (
                f"B{int(basement_to_floor.group(1))}-"
                f"{int(basement_to_floor.group(2))}F"
            )
        elif basement_range:
            readable_floors = (
                f"B{int(basement_range.group(1))}-"
                f"B{int(basement_range.group(2))}"
            )
        else:
            readable_floors = " + ".join(floors)

        floor_text = readable_floors

        if len(towers) == 1:
            review_location = (
                f"Cross-floor / {towers[0]} / {readable_floors}"
            )
        elif len(towers) > 1:
            review_location = (
                f"Distribution U / {', '.join(towers)} / "
                f"{readable_floors}"
            )
        else:
            review_location = f"Cross-floor / {readable_floors}"

        return (
            review_location,
            tower_text,
            floor_text,
            "跨樓層／多位置，人數分布未指定",
        )

    floor = floors[0]
    podium_floors = set(config.get("podium_floors", []))
    basement_floors = set(config.get("basement_floors", []))
    tower_floors = set(config.get("tower_floors", []))
    special_locations = set(config.get("special_locations", []))
    merge_podium = bool(config.get("merge_podium_tower_references", True))

    if floor in podium_floors and merge_podium:
        return f"Podium / {floor}", tower_text, floor_text, "已解析"

    if floor in basement_floors or floor.startswith("B"):
        return f"Basement / {floor}", tower_text, floor_text, "已解析"

    if floor in podium_floors:
        if towers:
            return f"{', '.join(towers)} / {floor}", tower_text, floor_text, "已解析"
        return f"Podium / {floor}", tower_text, floor_text, "已解析"

    if floor in tower_floors:
        if towers:
            return f"{', '.join(towers)} / {floor}", tower_text, floor_text, "已解析"
        return f"Unspecified Tower / {floor}", tower_text, floor_text, "需人工確認樓座"

    if floor in special_locations:
        return floor, tower_text, floor_text, "已解析"

    fallback = canonical_location or clean_cell(combined_text)
    return fallback, tower_text, floor_text, "需人工確認位置"


def extract_worker_table_rows(file_bytes: bytes) -> list[dict[str, str]]:
    """Extract rows from a Worker table containing Item/No., Name and Location."""
    worker_rows: list[dict[str, str]] = []

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []

            for table in tables:
                if not table:
                    continue

                cleaned_table = [
                    [clean_cell(cell) for cell in row]
                    for row in table
                    if row
                ]

                header_index = None
                item_index = None
                name_index = None
                location_index = None

                for row_index, row in enumerate(cleaned_table):
                    upper_row = [cell.upper() for cell in row]

                    name_candidates = [
                        index
                        for index, cell in enumerate(upper_row)
                        if cell == "NAME" or cell.endswith(" NAME")
                    ]
                    location_candidates = [
                        index
                        for index, cell in enumerate(upper_row)
                        if "LOCATION" in cell
                    ]

                    if not name_candidates or not location_candidates:
                        continue

                    header_index = row_index
                    name_index = name_candidates[0]
                    location_index = location_candidates[0]

                    item_candidates = [
                        index
                        for index, cell in enumerate(upper_row)
                        if cell in {"ITEM", "NO", "NO.", "NUMBER"}
                    ]
                    item_index = item_candidates[0] if item_candidates else 0
                    break

                if header_index is None:
                    continue

                for row in cleaned_table[header_index + 1 :]:
                    required_length = max(
                        item_index or 0,
                        name_index or 0,
                        location_index or 0,
                    ) + 1

                    if len(row) < required_length:
                        row += [""] * (required_length - len(row))

                    item = row[item_index or 0]
                    name = row[name_index or 0]
                    location = row[location_index or 0]

                    if not item and not name and not location:
                        continue

                    joined = " ".join(row).upper()
                    if "SUBMITTED BY" in joined or "SITE PHOTOS" in joined:
                        break

                    if name.upper() in {"NAME", "NIL", "NONE"}:
                        continue

                    # Worker rows normally use a plain integer item number.
                    if item and not re.fullmatch(r"\d+", item):
                        continue

                    if not name:
                        continue

                    worker_rows.append(
                        {
                            "item": item,
                            "name": name,
                            "location": location,
                        }
                    )

    return worker_rows


def build_worker_location_result(
    location_raw: str,
    config: dict[str, Any],
) -> tuple[str, str, str, str]:
    """Standardise worker-table locations without inventing a floor split."""
    original = clean_cell(location_raw)
    upper = original.upper().replace("＆", "&")

    if not original or upper in {"N/A", "NA", "NIL", "NONE", "ALL SITE", "SITE"}:
        return "Unspecified", "", "", "需人工確認位置"

    # Exact tower/floor or exact floor references can use the normal parser.
    floors = extract_floor_tokens(original)
    towers = extract_towers(original)

    if floors:
        return build_location_result(original, original, config)

    has_basement = "BASEMENT" in upper
    has_podium = "PODIUM" in upper
    tower_word_only = upper.strip() in {"TOWER", "TOWERS"}

    components: list[str] = []
    if has_basement:
        components.append("Basement")
    if has_podium:
        components.append("Podium")
    components.extend(towers)

    # Remove duplicates but keep a stable readable order.
    unique_components: list[str] = []
    for component in components:
        if component not in unique_components:
            unique_components.append(component)

    tower_text = ", ".join(towers)

    if tower_word_only:
        return "Unspecified Tower / Floor U", "", "", "需人工確認樓座及樓層"

    if len(unique_components) > 1:
        label = " + ".join(unique_components)
        return (
            f"Distribution U / {label}",
            tower_text,
            "",
            "位置分布未指定，不作假設分配",
        )

    if unique_components == ["Basement"]:
        return (
            "Basement / Floor U",
            "",
            "",
            "需人工確認樓層",
        )

    if unique_components == ["Podium"]:
        return (
            "Podium / Floor U",
            "",
            "",
            "需人工確認樓層",
        )

    if len(towers) == 1 and len(unique_components) == 1:
        return (
            f"{towers[0]} / Floor U",
            towers[0],
            "",
            "需人工確認樓層",
        )

    return original, tower_text, "", "需人工確認位置"


def analyse_worker_table_pdf(
    file_data: dict[str, Any],
    config: dict[str, Any],
    trade: str,
) -> dict[str, Any]:
    """Analyse FS or PD reports using the Worker attendance table."""
    raw_text = extract_pdf_text(file_data["bytes"])
    header = extract_header_data(raw_text)
    worker_rows = extract_worker_table_rows(file_data["bytes"])

    detail_rows: list[dict[str, Any]] = []

    for row in worker_rows:
        name = clean_cell(row.get("name", ""))
        location_raw = clean_cell(row.get("location", ""))
        standard_location, towers, floors, status = build_worker_location_result(
            location_raw,
            config,
        )

        detail_rows.append(
            {
                "日期": header["date"] or "",
                "工種": trade,
                "文件": file_data["name"],
                "Item": row.get("item", ""),
                "工作描述": "Worker attendance",
                "工人姓名": name,
                "原始位置": location_raw,
                "樓座": towers,
                "樓層": floors,
                "標準位置": standard_location,
                "人數": 1,
                "人數計算方法": "Worker表每列1人",
                "位置狀態": status,
            }
        )

    extracted_worker_total = len(detail_rows)
    worker = header["worker"]

    if worker is None:
        reconciliation = "未能讀取 Worker"
    elif extracted_worker_total == worker:
        reconciliation = "一致"
    else:
        reconciliation = f"不一致，相差 {worker - extracted_worker_total:+d}"

    names = [row["工人姓名"] for row in detail_rows if row["工人姓名"]]
    duplicate_names = sorted(
        {
            name
            for name in names
            if names.count(name) > 1
        }
    )

    summary = {
        "文件": file_data["name"],
        "日期": header["date"] or "未讀取",
        "工種": trade,
        "Ref. No.": header["reference"] or "未讀取",
        "Today Total Manpower": header["today_total"],
        "Worker": worker,
        "管理及技術人員": header["management_staff"],
        "工作明細人數合計": extracted_worker_total,
        "Worker核對": reconciliation,
        "明細提取方式": "Worker出勤表",
        "明細列數": len(detail_rows),
        "重複姓名": ", ".join(duplicate_names) if duplicate_names else "沒有",
    }

    return {
        "summary": summary,
        "details": detail_rows,
        "raw_text": raw_text,
    }


def analyse_el_pdf(
    file_data: dict[str, Any],
    config: dict[str, Any],
) -> dict[str, Any]:
    raw_text = extract_pdf_text(file_data["bytes"])
    header = extract_header_data(raw_text)
    site_rows = extract_site_work_table_rows(file_data["bytes"])

    extraction_method = "PDF表格"
    if not site_rows:
        site_rows = extract_fallback_detail_rows(raw_text)
        extraction_method = "文字後備規則"

    detail_rows: list[dict[str, Any]] = []

    for row in site_rows:
        description = clean_cell(row.get("description", ""))
        location_raw = clean_cell(row.get("location", ""))

        manpower, manpower_method = parse_description_manpower(description)
        standard_location, towers, floors, status = build_location_result(
            description,
            location_raw,
            config,
        )

        if manpower is None:
            status = "需人工確認人數"

        detail_rows.append(
            {
                "日期": header["date"] or "",
                "工種": "EL",
                "文件": file_data["name"],
                "Item": row.get("item", ""),
                "工作描述": description,
                "工人姓名": "",
                "原始位置": location_raw,
                "樓座": towers,
                "樓層": floors,
                "標準位置": standard_location,
                "人數": manpower,
                "人數計算方法": manpower_method,
                "位置狀態": status,
            }
        )

    extracted_worker_total = sum(
        int(row["人數"])
        for row in detail_rows
        if row["人數"] is not None
    )

    worker = header["worker"]
    if worker is None:
        reconciliation = "未能讀取 Worker"
    elif extracted_worker_total == worker:
        reconciliation = "一致"
    else:
        reconciliation = f"不一致，相差 {worker - extracted_worker_total:+d}"

    summary = {
        "文件": file_data["name"],
        "日期": header["date"] or "未讀取",
        "工種": header["trade"] or "EL",
        "Ref. No.": header["reference"] or "未讀取",
        "Today Total Manpower": header["today_total"],
        "Worker": worker,
        "管理及技術人員": header["management_staff"],
        "工作明細人數合計": extracted_worker_total,
        "Worker核對": reconciliation,
        "明細提取方式": extraction_method,
        "明細列數": len(detail_rows),
        "重複姓名": "不適用",
    }

    return {
        "summary": summary,
        "details": detail_rows,
        "raw_text": raw_text,
    }


# =========================================================
# Flexible multi-format manpower parsing
# =========================================================


def parse_integer_cell(value: Any) -> int | None:
    """Read a positive whole-number manpower value from one table cell."""
    text = clean_cell(value)
    if not text:
        return None

    match = re.fullmatch(r"\s*(\d+)\s*(?:人|PERSONS?|WORKERS?|PAX)?\s*", text, re.I)
    if not match:
        return None

    return int(match.group(1))


def extract_numeric_manpower_table_rows(
    file_bytes: bytes,
) -> list[dict[str, Any]]:
    """Extract rows from tables containing Location and Manpower-like columns."""
    extracted_rows: list[dict[str, Any]] = []

    manpower_keywords = (
        "MANPOWER",
        "NO. OF WORKER",
        "NO OF WORKER",
        "WORKER COUNT",
        "WORKERS",
        "PERSONS",
        "HEADCOUNT",
        "QTY",
        "QUANTITY",
    )
    location_keywords = (
        "LOCATION",
        "WORK LOCATION",
        "AREA",
        "ZONE",
        "FLOOR",
    )
    description_keywords = (
        "DESCRIPTION",
        "ACTIVITY",
        "WORK DESCRIPTION",
        "WORK ITEM",
    )

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                if not table:
                    continue

                cleaned_table = [
                    [clean_cell(cell) for cell in row]
                    for row in table
                    if row
                ]

                header_index = None
                manpower_index = None
                location_index = None
                description_index = None
                item_index = None

                for row_index, row in enumerate(cleaned_table):
                    upper_row = [cell.upper() for cell in row]

                    manpower_candidates = [
                        index
                        for index, cell in enumerate(upper_row)
                        if any(keyword in cell for keyword in manpower_keywords)
                    ]
                    location_candidates = [
                        index
                        for index, cell in enumerate(upper_row)
                        if any(keyword in cell for keyword in location_keywords)
                    ]

                    # A Location-like column is required so the general manpower
                    # summary table (Position / Persons) is not mistaken for
                    # location distribution data.
                    if not manpower_candidates or not location_candidates:
                        continue

                    header_index = row_index
                    manpower_index = manpower_candidates[0]
                    location_index = location_candidates[0]

                    description_candidates = [
                        index
                        for index, cell in enumerate(upper_row)
                        if any(keyword in cell for keyword in description_keywords)
                    ]
                    description_index = (
                        description_candidates[0]
                        if description_candidates
                        else None
                    )

                    item_candidates = [
                        index
                        for index, cell in enumerate(upper_row)
                        if cell in {"ITEM", "NO", "NO.", "NUMBER"}
                    ]
                    item_index = item_candidates[0] if item_candidates else None
                    break

                if header_index is None:
                    continue

                for row in cleaned_table[header_index + 1 :]:
                    required_indices = [
                        index
                        for index in (
                            manpower_index,
                            location_index,
                            description_index,
                            item_index,
                        )
                        if index is not None
                    ]
                    required_length = max(required_indices) + 1
                    if len(row) < required_length:
                        row += [""] * (required_length - len(row))

                    manpower = parse_integer_cell(row[manpower_index or 0])
                    if manpower is None or manpower <= 0:
                        continue

                    location = row[location_index or 0]
                    description = (
                        row[description_index]
                        if description_index is not None
                        else ""
                    )
                    item = row[item_index] if item_index is not None else ""

                    joined = " ".join(row).upper()
                    if any(word in joined for word in ("TOTAL", "SUBTOTAL", "GRAND TOTAL")):
                        continue

                    if not location and not description:
                        continue

                    extracted_rows.append(
                        {
                            "item": item,
                            "description": description,
                            "location": location,
                            "manpower": manpower,
                        }
                    )

    return extracted_rows


def build_standard_detail_row(
    *,
    header: dict[str, Any],
    file_data: dict[str, Any],
    config: dict[str, Any],
    trade: str,
    item: str,
    description: str,
    worker_name: str,
    location_raw: str,
    manpower: int | None,
    manpower_method: str,
    worker_location_mode: bool = False,
) -> dict[str, Any]:
    """Convert any extraction method into the common detail schema."""
    if worker_location_mode:
        standard_location, towers, floors, status = build_worker_location_result(
            location_raw,
            config,
        )
    else:
        standard_location, towers, floors, status = build_location_result(
            description,
            location_raw,
            config,
        )

    if manpower is None:
        status = "需人工確認人數"

    return {
        "日期": header["date"] or "",
        "工種": trade,
        "文件": file_data["name"],
        "Item": item,
        "工作描述": description,
        "工人姓名": worker_name,
        "原始位置": location_raw,
        "樓座": towers,
        "樓層": floors,
        "標準位置": standard_location,
        "人數": manpower,
        "人數計算方法": manpower_method,
        "位置狀態": status,
    }


def make_analysis_summary(
    *,
    header: dict[str, Any],
    file_data: dict[str, Any],
    trade: str,
    detail_rows: list[dict[str, Any]],
    extraction_method: str,
    duplicate_names: list[str] | None = None,
    reconciliation_override: str | None = None,
) -> dict[str, Any]:
    extracted_total = sum(
        int(row["人數"])
        for row in detail_rows
        if row.get("人數") is not None
    )
    worker = header["worker"]

    if reconciliation_override is not None:
        reconciliation = reconciliation_override
    elif worker is None:
        reconciliation = "未能讀取 Worker"
    elif extracted_total == worker:
        reconciliation = "一致"
    else:
        reconciliation = f"不一致，相差 {worker - extracted_total:+d}"

    return {
        "文件": file_data["name"],
        "日期": header["date"] or "未讀取",
        "工種": trade,
        "Ref. No.": header["reference"] or "未讀取",
        "Today Total Manpower": header["today_total"],
        "Worker": worker,
        "管理及技術人員": header["management_staff"],
        "工作明細人數合計": extracted_total,
        "Worker核對": reconciliation,
        "明細提取方式": extraction_method,
        "明細列數": len(detail_rows),
        "重複姓名": (
            ", ".join(duplicate_names)
            if duplicate_names
            else "沒有"
        ),
    }


def analyse_numeric_table_pdf(
    file_data: dict[str, Any],
    config: dict[str, Any],
    trade: str,
) -> dict[str, Any]:
    raw_text = extract_pdf_text(file_data["bytes"])
    header = extract_header_data(raw_text)
    numeric_rows = extract_numeric_manpower_table_rows(file_data["bytes"])

    detail_rows = [
        build_standard_detail_row(
            header=header,
            file_data=file_data,
            config=config,
            trade=trade,
            item=clean_cell(row.get("item", "")),
            description=clean_cell(row.get("description", "")),
            worker_name="",
            location_raw=clean_cell(row.get("location", "")),
            manpower=row.get("manpower"),
            manpower_method="Location＋Manpower數字欄",
        )
        for row in numeric_rows
    ]

    return {
        "summary": make_analysis_summary(
            header=header,
            file_data=file_data,
            trade=trade,
            detail_rows=detail_rows,
            extraction_method="Location＋Manpower數字欄",
        ),
        "details": detail_rows,
        "raw_text": raw_text,
    }


def analyse_description_pdf(
    file_data: dict[str, Any],
    config: dict[str, Any],
    trade: str,
) -> dict[str, Any]:
    raw_text = extract_pdf_text(file_data["bytes"])
    header = extract_header_data(raw_text)
    site_rows = extract_site_work_table_rows(file_data["bytes"])
    extraction_source = "PDF表格"

    if not site_rows:
        site_rows = extract_fallback_detail_rows(raw_text)
        extraction_source = "文字後備規則"

    detail_rows: list[dict[str, Any]] = []
    for row in site_rows:
        description = clean_cell(row.get("description", ""))
        location_raw = clean_cell(row.get("location", ""))
        manpower, method = parse_description_manpower(description)

        # Only keep rows where a count was actually found. This prevents
        # ordinary work-description tables from being treated as manpower
        # distribution tables.
        if manpower is None:
            continue

        detail_rows.append(
            build_standard_detail_row(
                header=header,
                file_data=file_data,
                config=config,
                trade=trade,
                item=clean_cell(row.get("item", "")),
                description=description,
                worker_name="",
                location_raw=location_raw,
                manpower=manpower,
                manpower_method=method,
            )
        )

    return {
        "summary": make_analysis_summary(
            header=header,
            file_data=file_data,
            trade=trade,
            detail_rows=detail_rows,
            extraction_method=f"{extraction_source}／描述人數",
        ),
        "details": detail_rows,
        "raw_text": raw_text,
    }


def analyse_total_only_pdf(
    file_data: dict[str, Any],
    config: dict[str, Any],
    trade: str,
    note: str = "",
) -> dict[str, Any]:
    raw_text = extract_pdf_text(file_data["bytes"])
    header = extract_header_data(raw_text)

    manpower = header["worker"]
    method = "報告Worker總數"
    if manpower is None:
        manpower = header["today_total"]
        method = "Today Total Manpower（可能包括管理人員）"

    detail_rows: list[dict[str, Any]] = []
    if manpower is not None:
        detail_rows.append(
            {
                "日期": header["date"] or "",
                "工種": trade,
                "文件": file_data["name"],
                "Item": "",
                "工作描述": "報告只提供總人數" + (f"；{note}" if note else ""),
                "工人姓名": "",
                "原始位置": "Unspecified",
                "樓座": "",
                "樓層": "",
                "標準位置": "Unspecified",
                "人數": manpower,
                "人數計算方法": method,
                "位置狀態": "需人工分配位置",
            }
        )

    return {
        "summary": make_analysis_summary(
            header=header,
            file_data=file_data,
            trade=trade,
            detail_rows=detail_rows,
            extraction_method="只使用報告總人數",
            reconciliation_override="位置未分配",
        ),
        "details": detail_rows,
        "raw_text": raw_text,
    }


def analyse_report_pdf(
    file_data: dict[str, Any],
    config: dict[str, Any],
    trade: str,
    parser_mode: str,
) -> dict[str, Any]:
    """Select the requested parser or auto-detect a suitable format."""
    if parser_mode == "numeric_table":
        result = analyse_numeric_table_pdf(file_data, config, trade)
        if result["details"]:
            return result
        return analyse_total_only_pdf(
            file_data,
            config,
            trade,
            "找不到Location＋Manpower數字表",
        )

    if parser_mode == "worker_table":
        result = analyse_worker_table_pdf(file_data, config, trade)
        if result["details"]:
            return result
        return analyse_total_only_pdf(
            file_data,
            config,
            trade,
            "找不到工人姓名表",
        )

    if parser_mode == "description":
        result = analyse_description_pdf(file_data, config, trade)
        if result["details"]:
            return result
        return analyse_total_only_pdf(
            file_data,
            config,
            trade,
            "工作描述中找不到可計算人數",
        )

    if parser_mode == "total_only":
        return analyse_total_only_pdf(file_data, config, trade)

    # Automatic detection. Prefer explicit location/manpower tables. For FS
    # and PD, a worker attendance table is normally more reliable than any
    # incidental number appearing in the work-description section.
    numeric_result = analyse_numeric_table_pdf(file_data, config, trade)
    if numeric_result["details"]:
        numeric_result["summary"]["明細提取方式"] = "自動偵測：Location＋Manpower數字欄"
        return numeric_result

    if trade in {"FS", "PD"}:
        worker_result = analyse_worker_table_pdf(file_data, config, trade)
        if worker_result["details"]:
            worker_result["summary"]["明細提取方式"] = "自動偵測：工人姓名表"
            return worker_result

        description_result = analyse_description_pdf(file_data, config, trade)
        if description_result["details"]:
            description_result["summary"]["明細提取方式"] = "自動偵測：工作描述人數"
            return description_result
    else:
        description_result = analyse_description_pdf(file_data, config, trade)
        if description_result["details"]:
            description_result["summary"]["明細提取方式"] = "自動偵測：工作描述人數"
            return description_result

        worker_result = analyse_worker_table_pdf(file_data, config, trade)
        if worker_result["details"]:
            worker_result["summary"]["明細提取方式"] = "自動偵測：工人姓名表"
            return worker_result

    return analyse_total_only_pdf(
        file_data,
        config,
        trade,
        "自動偵測不到位置人數明細",
    )




# =========================================================
# Excel template export functions
# =========================================================


# =========================================================
# Built-in blank workbook template
# =========================================================


MASTER_HEADER_FILL = "1F4E78"
MASTER_HEADER_FONT = "FFFFFF"
MASTER_LIGHT_BLUE_FILL = "D9EAF7"
MASTER_DARK_BLUE_FONT = "203864"
MASTER_YELLOW_FILL = "FFF2CC"
MASTER_GRID_COLOR = "D9E2F3"


def safe_filename_part(value: str) -> str:
    """Return a Windows/macOS-safe filename component."""
    cleaned = re.sub(
        r'[\\/:*?"<>|]+',
        "_",
        clean_cell(value),
    )
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    return cleaned or "New_Project"


def set_master_header_style(
    worksheet: Any,
    cell_range: str,
) -> None:
    """Apply the workbook's dark-blue table header style."""
    header_fill = PatternFill(
        fill_type="solid",
        fgColor=MASTER_HEADER_FILL,
    )
    header_font = Font(
        name="Times New Roman",
        size=12,
        bold=True,
        color=MASTER_HEADER_FONT,
    )
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    thin_side = Side(
        style="thin",
        color=MASTER_GRID_COLOR,
    )
    header_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    for row in worksheet[cell_range]:
        for cell in row:
            cell.fill = copy(header_fill)
            cell.font = copy(header_font)
            cell.alignment = copy(header_alignment)
            cell.border = copy(header_border)


def set_master_body_style(
    worksheet: Any,
    cell_range: str,
    *,
    yellow: bool = False,
) -> None:
    """Apply the standard Times New Roman body style."""
    body_font = Font(
        name="Times New Roman",
        size=12,
        color="000000",
    )
    body_fill = PatternFill(
        fill_type="solid",
        fgColor=(
            MASTER_YELLOW_FILL
            if yellow
            else "FFFFFF"
        ),
    )
    thin_side = Side(
        style="thin",
        color=MASTER_GRID_COLOR,
    )
    body_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    for row in worksheet[cell_range]:
        for cell in row:
            cell.font = copy(body_font)
            cell.fill = copy(body_fill)
            cell.border = copy(body_border)
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(
                    cell.column in {3, 9}
                ),
            )


def add_master_table(
    worksheet: Any,
    reference: str,
    table_name: str,
) -> None:
    """Add a blue banded Excel table matching the current workbook."""
    table = Table(
        displayName=table_name,
        ref=reference,
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def create_blank_master_template_bytes(
    config: dict[str, Any],
) -> bytes:
    """Create a completely new blank workbook in the current master layout."""
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    project_name = clean_cell(
        config.get("project_name")
    ) or "New Project"

    # -----------------------------------------------------
    # Overview
    # -----------------------------------------------------
    overview = workbook.create_sheet("Overview")
    overview.merge_cells("A1:D1")
    overview["A1"] = f"{project_name} Manpower Overview"
    overview["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor=MASTER_HEADER_FILL,
    )
    overview["A1"].font = Font(
        name="Times New Roman",
        size=16,
        bold=True,
        color=MASTER_HEADER_FONT,
    )
    overview["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    overview.row_dimensions[1].height = 28

    overview["A4"] = "Metric"
    overview["B4"] = "Value"
    overview["D4"] = "How to use the location analysis"
    set_master_header_style(overview, "A4:B4")
    set_master_header_style(overview, "D4:D4")

    overview_rows = [
        (
            "Start Date",
            '=IFERROR(MIN(\'Daily Master\'!A:A),"")',
        ),
        (
            "Last Updated",
            '=IFERROR(MAX(\'Daily Master\'!A:A),"")',
        ),
        (
            "Total Manpower",
            "='Department Summary'!B6",
        ),
        ("AC", "='Department Summary'!B2"),
        ("EL", "='Department Summary'!B3"),
        ("FS", "='Department Summary'!B4"),
        ("PD", "='Department Summary'!B5"),
        (
            "Days with Reports",
            '=COUNT(\'Daily Master\'!A:A)',
        ),
    ]
    for row_number, (metric, formula) in enumerate(
        overview_rows,
        start=5,
    ):
        overview.cell(row_number, 1).value = metric
        overview.cell(row_number, 2).value = formula

    overview["D5"] = (
        "Use “Location Detail” to filter by date, trade, tower or "
        "work location. Cross-floor and multi-location records are "
        "kept once in “Cross-F & distribution U” when the reports do "
        "not state the exact worker distribution."
    )
    overview.merge_cells("D5:D12")

    set_master_body_style(overview, "A5:B12")
    set_master_body_style(overview, "D5:D12")
    overview["D5"].alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )
    overview["B5"].number_format = "yyyy-mm-dd"
    overview["B6"].number_format = "yyyy-mm-dd"
    overview["A12"].fill = PatternFill(
        fill_type="solid",
        fgColor=MASTER_LIGHT_BLUE_FILL,
    )
    overview["B12"].fill = PatternFill(
        fill_type="solid",
        fgColor=MASTER_LIGHT_BLUE_FILL,
    )
    overview["A12"].font = Font(
        name="Times New Roman",
        size=12,
        bold=True,
        color=MASTER_DARK_BLUE_FONT,
    )
    overview["B12"].font = copy(overview["A12"].font)
    overview.column_dimensions["A"].width = 23
    overview.column_dimensions["B"].width = 18
    overview.column_dimensions["C"].width = 3
    overview.column_dimensions["D"].width = 58
    overview.freeze_panes = "A4"

    # -----------------------------------------------------
    # Data Rules
    # -----------------------------------------------------
    rules = workbook.create_sheet("Data Rules")
    rules.append(
        [
            "Rule / Item",
            "Definition used in this workbook",
        ]
    )

    basement_text = ", ".join(
        config.get("basement_floors", [])
    ) or "No basement floors configured"
    towers_text = ", ".join(
        config.get("towers", [])
    ) or "No towers configured"

    rule_rows = [
        [
            "Master manpower count",
            (
                "Taken from Section 'Today Total Manpower' of each "
                "daily report. Confirmed duplicate reports are counted once."
            ),
        ],
        [
            "Main Location Detail",
            (
                "Contains only single-location records or single "
                "tower/basement records where the exact floor may be unspecified."
            ),
        ],
        [
            "Cross-floor & Unspecified",
            (
                "Contains all multi-location / cross-floor worker records. "
                "Workers are retained once and are not automatically split."
            ),
        ],
        [
            "Exact multi-location",
            (
                "A worker explicitly listed at multiple exact locations is "
                "counted once in the Cross-floor sheet."
            ),
        ],
        [
            "Tower only",
            (
                "A tower without a floor is shown in Location Detail as "
                "floor unspecified and highlighted."
            ),
        ],
        [
            "Basement only",
            (
                "Basement without an exact floor is shown as floor "
                "unspecified and highlighted."
            ),
        ],
        [
            "Podium / Basement",
            (
                "Kept in Cross-floor & distribution U because the exact "
                "worker distribution is unknown."
            ),
        ],
        [
            "Project structure",
            f"Towers: {towers_text}. Basement floors: {basement_text}.",
        ],
        [
            "Highlighting",
            (
                "Pale yellow is used when the exact floor/location or "
                "distribution is genuinely uncertain."
            ),
        ],
    ]
    for row in rule_rows:
        rules.append(row)

    set_master_header_style(rules, "A1:B1")
    set_master_body_style(
        rules,
        f"A2:B{len(rule_rows) + 1}",
    )
    rules.column_dimensions["A"].width = 27
    rules.column_dimensions["B"].width = 105
    for row_number in range(2, len(rule_rows) + 2):
        rules.cell(row_number, 2).alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )
        rules.row_dimensions[row_number].height = 38
    rules.freeze_panes = "A2"

    # -----------------------------------------------------
    # Daily Master
    # -----------------------------------------------------
    daily = workbook.create_sheet("Daily Master")
    daily.append(
        [
            "Date",
            "AC",
            "EL",
            "FS",
            "PD",
            "Daily Total",
        ]
    )
    daily.append([None, None, None, None, None, "=SUM(B2:E2)"])
    set_master_header_style(daily, "A1:F1")
    set_master_body_style(daily, "A2:F2")
    daily["A2"].number_format = "yyyy-mm-dd"
    for column in "BCDEF":
        daily[f"{column}2"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
    add_master_table(
        daily,
        "A1:F2",
        "DailyMasterTable",
    )
    daily.freeze_panes = "A2"
    daily.column_dimensions["A"].width = 15
    for column in "BCDE":
        daily.column_dimensions[column].width = 11
    daily.column_dimensions["F"].width = 15

    # -----------------------------------------------------
    # Location sheets
    # -----------------------------------------------------
    location_headers = [
        "Date",
        "Area / Tower",
        "Location",
        "AC",
        "EL",
        "FS",
        "PD",
        "Allocated Site Workers",
        "Allocation Note",
    ]

    for (
        sheet_name,
        table_name,
    ) in (
        (
            "Location Detail",
            "LocationDetailTable",
        ),
        (
            "Cross-F & distribution U",
            "CrossFloorUnspecifiedTable",
        ),
    ):
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(location_headers)
        worksheet.append(
            [
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "=SUM(D2:G2)",
                None,
            ]
        )
        set_master_header_style(worksheet, "A1:I1")
        set_master_body_style(worksheet, "A2:I2")
        worksheet["A2"].number_format = "yyyy-mm-dd"
        for column in "DEFGH":
            worksheet[f"{column}2"].alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
        worksheet["C2"].alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )
        worksheet["I2"].alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )
        add_master_table(
            worksheet,
            "A1:I2",
            table_name,
        )
        worksheet.freeze_panes = "A2"
        widths = {
            "A": 15,
            "B": 30,
            "C": 52,
            "D": 9,
            "E": 9,
            "F": 9,
            "G": 9,
            "H": 22,
            "I": 76,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[
                column
            ].width = width
        worksheet.row_dimensions[2].height = 34

    # -----------------------------------------------------
    # Department Summary
    # -----------------------------------------------------
    department = workbook.create_sheet(
        "Department Summary"
    )
    department.append(
        ["Department", "Total Manpower"]
    )
    department_rows = [
        ["AC", "=SUM('Daily Master'!B:B)"],
        ["EL", "=SUM('Daily Master'!C:C)"],
        ["FS", "=SUM('Daily Master'!D:D)"],
        ["PD", "=SUM('Daily Master'!E:E)"],
        ["Grand Total", "=SUM(B2:B5)"],
    ]
    for row in department_rows:
        department.append(row)

    set_master_header_style(department, "A1:B1")
    set_master_body_style(department, "A2:B6")
    for cell in department["B"]:
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
    for column in "AB":
        department[f"{column}6"].fill = PatternFill(
            fill_type="solid",
            fgColor=MASTER_LIGHT_BLUE_FILL,
        )
        department[f"{column}6"].font = Font(
            name="Times New Roman",
            size=12,
            bold=True,
            color=MASTER_DARK_BLUE_FONT,
        )
    department.column_dimensions["A"].width = 24
    department.column_dimensions["B"].width = 20

    # Put the workbook in a normal calculation state.
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except Exception:
        pass

    output_buffer = io.BytesIO()
    workbook.save(output_buffer)
    return output_buffer.getvalue()




EXCEL_REQUIRED_SHEETS = (
    "Overview",
    "Data Rules",
    "Daily Master",
    "Location Detail",
    "Cross-F & distribution U",
    "Department Summary",
)

TRADE_COLUMN_MAP = {
    "AC": 4,
    "EL": 5,
    "FS": 6,
    "PD": 7,
}

DAILY_TRADE_COLUMN_MAP = {
    "AC": 2,
    "EL": 3,
    "FS": 4,
    "PD": 5,
}


def parse_date_value(value: Any) -> date | None:
    """Convert report, pandas or Excel date values into a Python date."""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = clean_cell(value)
    if not text or text in {"未讀取", "分析失敗"}:
        return None

    parsed = pd.to_datetime(text, errors="coerce", dayfirst=False)
    if pd.isna(parsed):
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)

    if pd.isna(parsed):
        return None

    return parsed.date()


def to_non_negative_int(value: Any) -> int | None:
    """Convert a cell-like value into a non-negative integer."""
    if value is None or value == "":
        return None

    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None

    if pd.isna(numeric) or numeric < 0:
        return None

    return int(round(numeric))


def build_daily_updates(
    summary_rows: list[dict[str, Any]],
) -> tuple[
    dict[date, dict[str, int]],
    dict[date, set[str]],
    list[str],
]:
    """Prepare Daily Master values and the date/trade update scope."""
    updates: dict[date, dict[str, int]] = {}
    update_scope: dict[date, set[str]] = {}
    seen_keys: set[tuple[date, str]] = set()
    duplicate_keys: list[str] = []

    for summary in summary_rows:
        report_date = parse_date_value(summary.get("日期"))
        trade = clean_cell(summary.get("工種")).upper()

        if report_date is None or trade not in REQUIRED_TRADES:
            continue

        update_scope.setdefault(report_date, set()).add(trade)

        # The user's workbook Data Rules define Daily Master as
        # Today Total Manpower, not Worker.
        daily_total = to_non_negative_int(
            summary.get("Today Total Manpower")
        )

        if daily_total is None:
            daily_total = to_non_negative_int(summary.get("Worker"))

        if daily_total is None:
            daily_total = to_non_negative_int(
                summary.get("工作明細人數合計")
            )

        if daily_total is None:
            continue

        key = (report_date, trade)
        if key in seen_keys:
            duplicate_keys.append(
                f"{report_date.isoformat()} {trade}"
            )

        seen_keys.add(key)
        updates.setdefault(report_date, {})[trade] = daily_total

    return updates, update_scope, sorted(set(duplicate_keys))


def standard_location_to_workbook(
    standard_location: str,
) -> tuple[str, str, str, bool]:
    """Convert an app location into the workbook fields and note."""
    location = clean_cell(standard_location)

    tower_exact = re.fullmatch(
        r"T\s*(\d+)\s*/\s*(\d+F)",
        location,
        re.I,
    )
    if tower_exact:
        tower_number = int(tower_exact.group(1))
        floor = tower_exact.group(2).upper()
        return (
            f"Tower {tower_number}",
            f"T{tower_number} {floor}",
            "Exact/specific work location recorded in the daily report.",
            False,
        )

    podium_exact = re.fullmatch(
        r"Podium\s*/\s*(GF|\d+F)",
        location,
        re.I,
    )
    if podium_exact:
        return (
            "Podium",
            podium_exact.group(1).upper(),
            "Exact/specific work location recorded in the daily report.",
            False,
        )

    basement_exact = re.fullmatch(
        r"Basement\s*/\s*(B\d+)",
        location,
        re.I,
    )
    if basement_exact:
        return (
            "Basement",
            basement_exact.group(1).upper(),
            "Exact/specific work location recorded in the daily report.",
            False,
        )

    tower_floor_unspecified = re.fullmatch(
        r"T\s*(\d+)\s*/\s*Floor\s*U",
        location,
        re.I,
    )
    if tower_floor_unspecified:
        tower_number = int(tower_floor_unspecified.group(1))
        return (
            f"Tower {tower_number}",
            f"T{tower_number} (floor unspecified)",
            "Tower stated, but the exact floor was not stated in the daily report.",
            True,
        )

    if re.fullmatch(r"Podium\s*/\s*Floor\s*U", location, re.I):
        return (
            "Podium",
            "Podium (floor unspecified)",
            "Podium stated, but the exact floor was not stated in the daily report.",
            True,
        )

    if re.fullmatch(r"Basement\s*/\s*Floor\s*U", location, re.I):
        return (
            "Basement",
            "Basement (floor unspecified)",
            "Basement stated, but the exact floor was not stated in the daily report.",
            True,
        )

    unspecified_tower_floor = re.fullmatch(
        r"Unspecified Tower\s*/\s*(GF|\d+F)",
        location,
        re.I,
    )
    if unspecified_tower_floor:
        return (
            "Tower (unspecified)",
            unspecified_tower_floor.group(1).upper(),
            "Floor stated, but the exact tower was not stated in the daily report.",
            True,
        )

    if re.fullmatch(
        r"Unspecified Tower\s*/\s*Floor\s*U",
        location,
        re.I,
    ):
        return (
            "Tower",
            "Tower (tower/floor unspecified)",
            "Tower area stated, but the exact tower and floor were not stated in the daily report.",
            True,
        )

    special_location = re.fullmatch(
        r"(Roof|M/F|UG/F|Clubhouse)",
        location,
        re.I,
    )
    if special_location:
        special = special_location.group(1)
        return (
            "Special",
            special,
            "Exact/specific work location recorded in the daily report.",
            False,
        )

    return (
        "Other",
        location or "Location unspecified",
        "The exact work location was not stated in the daily report.",
        True,
    )

def is_single_area_uncertain_location(
    standard_location: str,
) -> bool:
    """Return True when an uncertain item still belongs in Location Detail."""
    location = clean_cell(standard_location)

    patterns = (
        r"T\s*\d+\s*/\s*Floor\s*U",
        r"Podium\s*/\s*Floor\s*U",
        r"Basement\s*/\s*Floor\s*U",
        r"Unspecified Tower\s*/\s*(?:GF|\d+F)",
        r"Unspecified Tower\s*/\s*Floor\s*U",
    )

    return any(
        re.fullmatch(pattern, location, re.I)
        for pattern in patterns
    )


def format_cross_floor_location(
    standard_location: str,
) -> tuple[str, str, bool, str]:
    """Create Cross-F workbook text. All such rows are uncertain."""
    location = clean_cell(standard_location)
    area = "Cross-floor / Multi-location"

    if not location or location.lower() == "unspecified":
        return (
            area,
            "Location not stated (distribution unspecified)",
            True,
            "No work location was stated in the daily report; worker distribution is unspecified.",
        )

    tower_cross = re.fullmatch(
        r"Cross-floor\s*/\s*T\s*(\d+)\s*/\s*(.+)",
        location,
        re.I,
    )
    if tower_cross:
        tower = f"T{int(tower_cross.group(1))}"
        floor_text = clean_cell(tower_cross.group(2))

        # Keep compact ranges such as B4-3F intact; otherwise prefix the
        # tower to each listed floor.
        if re.fullmatch(r"B\d+-\d+F", floor_text, re.I):
            formatted = floor_text.upper()
        else:
            floors = [
                clean_cell(item).upper()
                for item in floor_text.split("+")
                if clean_cell(item)
            ]
            formatted = " + ".join(
                f"{tower} {floor}"
                for floor in floors
            )

        if not formatted.lower().endswith(
            "(distribution unspecified)"
        ):
            formatted += " (distribution unspecified)"

        return (
            area,
            formatted,
            True,
            "Multiple floors/locations were stated, but worker distribution between them was not specified.",
        )

    general_cross = re.fullmatch(
        r"Cross-floor\s*/\s*(.+)",
        location,
        re.I,
    )
    if general_cross:
        formatted = clean_cell(general_cross.group(1))
        if not formatted.lower().endswith(
            "(distribution unspecified)"
        ):
            formatted += " (distribution unspecified)"

        return (
            area,
            formatted,
            True,
            "Multiple floors/locations were stated, but worker distribution between them was not specified.",
        )

    distribution = re.fullmatch(
        r"Distribution U\s*/\s*(.+)",
        location,
        re.I,
    )
    if distribution:
        formatted = clean_cell(distribution.group(1))
        if not formatted.lower().endswith(
            "(distribution unspecified)"
        ):
            formatted += " (distribution unspecified)"

        return (
            area,
            formatted,
            True,
            "Multiple areas/towers were stated, but worker distribution between them was not specified.",
        )

    formatted = location
    if not formatted.lower().endswith(
        "(distribution unspecified)"
    ):
        formatted += " (distribution unspecified)"

    return (
        area,
        formatted,
        True,
        "The exact worker distribution was not stated in the daily report.",
    )

def workbook_record_sort_key(
    record: dict[str, Any],
) -> tuple[Any, ...]:
    """Sort only the updated date block in practical site order."""
    report_date = record["date"]
    area = clean_cell(record["area"])
    location = clean_cell(record["location"])

    tower_area = re.fullmatch(r"Tower\s+(\d+)", area, re.I)
    tower_exact = re.fullmatch(
        r"T\s*(\d+)\s+(\d+)F",
        location,
        re.I,
    )
    if tower_area and tower_exact:
        tower_number = int(tower_area.group(1))
        floor_number = int(tower_exact.group(2))
        return (report_date, 0, tower_number, -floor_number, location)

    tower_unspecified = re.fullmatch(
        r"T\s*(\d+)\s*\(floor unspecified\)",
        location,
        re.I,
    )
    if tower_area and tower_unspecified:
        tower_number = int(tower_area.group(1))
        return (report_date, 0, tower_number, 9999, location)

    if area.lower() == "tower" and re.fullmatch(
        r"Tower\s*\(tower/floor unspecified\)",
        location,
        re.I,
    ):
        return (report_date, 0, 999, 9999, location)

    if area.lower() == "podium":
        podium_floor = re.fullmatch(r"(GF|\d+F)", location, re.I)
        if podium_floor:
            floor = podium_floor.group(1).upper()
            floor_rank = 0 if floor == "GF" else -int(
                re.search(r"\d+", floor).group()
            )
            return (report_date, 1, 0, floor_rank, location)

        if "floor unspecified" in location.lower():
            return (report_date, 1, 0, 9999, location)

    if area.lower() == "basement":
        basement_floor = re.fullmatch(r"B(\d+)", location, re.I)
        if basement_floor:
            return (
                report_date,
                2,
                0,
                int(basement_floor.group(1)),
                location,
            )

        if "floor unspecified" in location.lower():
            return (report_date, 2, 0, 9999, location)

    return (report_date, 3, 0, 0, f"{area} {location}")

def read_location_records(
    worksheet: Any,
) -> list[dict[str, Any]]:
    """Read populated records from Location Detail or Cross-F."""
    records: list[dict[str, Any]] = []

    for row_number in range(2, worksheet.max_row + 1):
        report_date = parse_date_value(
            worksheet.cell(row_number, 1).value
        )
        area = clean_cell(
            worksheet.cell(row_number, 2).value
        )
        location = clean_cell(
            worksheet.cell(row_number, 3).value
        )

        if report_date is None or not area or not location:
            continue

        values = {
            trade: (
                to_non_negative_int(
                    worksheet.cell(
                        row_number,
                        TRADE_COLUMN_MAP[trade],
                    ).value
                )
                or 0
            )
            for trade in REQUIRED_TRADES
        }

        records.append(
            {
                "date": report_date,
                "area": area,
                "location": location,
                **values,
                "note": clean_cell(
                    worksheet.cell(row_number, 9).value
                ),
                "uncertain": (
                    "unspecified"
                    in clean_cell(
                        worksheet.cell(row_number, 9).value
                    ).lower()
                ),
            }
        )

    return records


def merge_location_updates(
    existing_records: list[dict[str, Any]],
    new_records: list[dict[str, Any]],
    update_scope: dict[date, set[str]],
) -> list[dict[str, Any]]:
    """Replace only the analysed date/trade values, preserving other trades."""
    record_map: dict[
        tuple[date, str, str],
        dict[str, Any],
    ] = {}

    for record in existing_records:
        copied_record = dict(record)

        for trade in update_scope.get(
            copied_record["date"],
            set(),
        ):
            copied_record[trade] = 0

        key = (
            copied_record["date"],
            copied_record["area"],
            copied_record["location"],
        )
        record_map[key] = copied_record

    for new_record in new_records:
        key = (
            new_record["date"],
            new_record["area"],
            new_record["location"],
        )

        target = record_map.setdefault(
            key,
            {
                "date": new_record["date"],
                "area": new_record["area"],
                "location": new_record["location"],
                "AC": 0,
                "EL": 0,
                "FS": 0,
                "PD": 0,
                "note": new_record["note"],
                "uncertain": new_record["uncertain"],
            },
        )

        for trade in update_scope.get(
            new_record["date"],
            set(),
        ):
            if trade in new_record:
                target[trade] = (
                    to_non_negative_int(
                        new_record.get(trade)
                    )
                    or 0
                )

        target["note"] = new_record["note"]
        target["uncertain"] = new_record["uncertain"]

    return [
        record
        for record in record_map.values()
        if sum(
            int(record.get(trade, 0) or 0)
            for trade in REQUIRED_TRADES
        )
        > 0
    ]


def find_sample_row(
    worksheet: Any,
    *,
    exact: bool,
    fallback: int = 2,
) -> int:
    """Find an existing exact or yellow uncertain row for style copying."""
    for row_number in range(2, worksheet.max_row + 1):
        note = clean_cell(
            worksheet.cell(row_number, 9).value
        ).lower()

        if not note:
            continue

        is_uncertain = (
            "unspecified" in note
            or "distribution" in note
        )

        if exact and not is_uncertain:
            return row_number

        if not exact and is_uncertain:
            return row_number

    return fallback


def copy_row_style(
    worksheet: Any,
    source_row: int,
    target_row: int,
    max_column: int,
) -> None:
    """Copy cell styles and row height without copying cell values."""
    for column_number in range(1, max_column + 1):
        source_cell = worksheet.cell(
            source_row,
            column_number,
        )
        target_cell = worksheet.cell(
            target_row,
            column_number,
        )
        target_cell._style = copy(source_cell._style)
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.number_format = source_cell.number_format

    worksheet.row_dimensions[target_row].height = (
        worksheet.row_dimensions[source_row].height
    )


def resize_table(
    worksheet: Any,
    table_name: str,
    reference: str,
) -> None:
    """Expand a named table when required, but never shrink the template."""
    if table_name not in worksheet.tables:
        return

    table: Table = worksheet.tables[table_name]
    old_min_col, old_min_row, old_max_col, old_max_row = (
        range_boundaries(table.ref)
    )
    new_min_col, new_min_row, new_max_col, new_max_row = (
        range_boundaries(reference)
    )

    min_col = min(old_min_col, new_min_col)
    min_row = min(old_min_row, new_min_row)
    max_col = max(old_max_col, new_max_col)
    max_row = max(old_max_row, new_max_row)

    table.ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )

def rewrite_location_sheet(
    worksheet: Any,
    records: list[dict[str, Any]],
    *,
    table_name: str,
    cross_sheet: bool,
) -> None:
    """Rewrite one location sheet while preserving its template formatting."""
    exact_style_row = find_sample_row(
        worksheet,
        exact=True,
        fallback=2,
    )
    uncertain_style_row = find_sample_row(
        worksheet,
        exact=False,
        fallback=exact_style_row,
    )

    existing_max_row = worksheet.max_row
    required_last_row = max(2, len(records) + 1)
    clear_last_row = max(
        existing_max_row,
        required_last_row,
    )

    for row_number in range(2, clear_last_row + 1):
        for column_number in range(1, 10):
            worksheet.cell(
                row_number,
                column_number,
            ).value = None

    for index, record in enumerate(records, start=2):
        style_row = (
            uncertain_style_row
            if record.get("uncertain")
            else exact_style_row
        )
        copy_row_style(
            worksheet,
            style_row,
            index,
            9,
        )

        worksheet.cell(index, 1).value = record["date"]
        worksheet.cell(index, 1).number_format = "yyyy-mm-dd"
        worksheet.cell(index, 2).value = record["area"]
        worksheet.cell(index, 3).value = record["location"]

        for trade in REQUIRED_TRADES:
            worksheet.cell(
                index,
                TRADE_COLUMN_MAP[trade],
            ).value = int(
                record.get(trade, 0) or 0
            )

        worksheet.cell(index, 8).value = (
            f"=SUM(D{index}:G{index})"
        )
        worksheet.cell(index, 9).value = record["note"]

    resize_table(
        worksheet,
        table_name,
        f"A1:I{required_last_row}",
    )


def is_reusable_blank_location_row(
    worksheet: Any,
    row_number: int,
) -> bool:
    """Return True for an unused/template data row."""
    date_value = parse_date_value(
        worksheet.cell(row_number, 1).value
    )
    area = clean_cell(worksheet.cell(row_number, 2).value)
    location = clean_cell(worksheet.cell(row_number, 3).value)
    note = clean_cell(worksheet.cell(row_number, 9).value)

    trade_total = sum(
        to_non_negative_int(
            worksheet.cell(
                row_number,
                TRADE_COLUMN_MAP[trade],
            ).value
        )
        or 0
        for trade in REQUIRED_TRADES
    )

    return (
        date_value is None
        and not area
        and not location
        and trade_total == 0
        and not note
    )


def read_location_records_for_date(
    worksheet: Any,
    report_date: date,
    row_numbers: list[int],
) -> list[dict[str, Any]]:
    """Read populated rows for one date without touching other dates."""
    records: list[dict[str, Any]] = []

    for row_number in row_numbers:
        if parse_date_value(
            worksheet.cell(row_number, 1).value
        ) != report_date:
            continue

        area = clean_cell(
            worksheet.cell(row_number, 2).value
        )
        location = clean_cell(
            worksheet.cell(row_number, 3).value
        )

        if not area or not location:
            continue

        values = {
            trade: (
                to_non_negative_int(
                    worksheet.cell(
                        row_number,
                        TRADE_COLUMN_MAP[trade],
                    ).value
                )
                or 0
            )
            for trade in REQUIRED_TRADES
        }

        records.append(
            {
                "date": report_date,
                "area": area,
                "location": location,
                **values,
                "note": clean_cell(
                    worksheet.cell(row_number, 9).value
                ),
                "uncertain": bool(
                    worksheet.cell(row_number, 1).fill
                    and worksheet.cell(row_number, 1).fill.fill_type
                    and worksheet.cell(row_number, 1).fill.fgColor.rgb
                    not in {None, "00000000", "FFFFFFFF"}
                )
                or "unspecified"
                in clean_cell(
                    worksheet.cell(row_number, 9).value
                ).lower()
                or "distribution"
                in clean_cell(
                    worksheet.cell(row_number, 9).value
                ).lower(),
            }
        )

    return records


def find_location_insertion_row(
    worksheet: Any,
    report_date: date,
    scan_last_row: int,
) -> int:
    """Find where a new date block should be written."""
    last_dated_row = 1

    for row_number in range(2, scan_last_row + 1):
        existing_date = parse_date_value(
            worksheet.cell(row_number, 1).value
        )

        if existing_date is None:
            continue

        if existing_date > report_date:
            return row_number

        last_dated_row = row_number

    return last_dated_row + 1


def clear_location_row_values(
    worksheet: Any,
    row_number: int,
) -> None:
    """Clear A:I values while retaining the row's formatting."""
    for column_number in range(1, 10):
        worksheet.cell(
            row_number,
            column_number,
        ).value = None


def write_location_record(
    worksheet: Any,
    row_number: int,
    record: dict[str, Any],
    exact_style_row: int,
    uncertain_style_row: int,
) -> None:
    """Write one location record using the matching template style."""
    style_row = (
        uncertain_style_row
        if record.get("uncertain")
        else exact_style_row
    )
    copy_row_style(worksheet, style_row, row_number, 9)

    worksheet.cell(row_number, 1).value = record["date"]
    worksheet.cell(row_number, 1).number_format = "yyyy-mm-dd"
    worksheet.cell(row_number, 2).value = record["area"]
    worksheet.cell(row_number, 3).value = record["location"]

    for trade in REQUIRED_TRADES:
        worksheet.cell(
            row_number,
            TRADE_COLUMN_MAP[trade],
        ).value = int(record.get(trade, 0) or 0)

    worksheet.cell(row_number, 8).value = (
        f"=SUM(D{row_number}:G{row_number})"
    )
    worksheet.cell(row_number, 9).value = record["note"]

    if record.get("uncertain"):
        yellow_fill = PatternFill(
            fill_type="solid",
            fgColor="FFF2CC",
        )
        for column_number in range(1, 10):
            worksheet.cell(
                row_number,
                column_number,
            ).fill = copy(yellow_fill)


def update_location_sheet_preserving_history(
    worksheet: Any,
    new_records: list[dict[str, Any]],
    update_scope: dict[date, set[str]],
    *,
    table_name: str,
    cross_sheet: bool,
) -> int:
    """Update only affected date blocks; leave all older rows untouched."""
    exact_style_row = find_sample_row(
        worksheet,
        exact=True,
        fallback=2,
    )
    uncertain_style_row = find_sample_row(
        worksheet,
        exact=False,
        fallback=exact_style_row,
    )

    if table_name in worksheet.tables:
        _, _, _, table_last_row = range_boundaries(
            worksheet.tables[table_name].ref
        )
    else:
        table_last_row = worksheet.max_row

    for report_date in sorted(update_scope):
        date_rows = [
            row_number
            for row_number in range(2, table_last_row + 1)
            if parse_date_value(
                worksheet.cell(row_number, 1).value
            )
            == report_date
        ]

        existing_records = read_location_records_for_date(
            worksheet,
            report_date,
            date_rows,
        )
        date_new_records = [
            record
            for record in new_records
            if record["date"] == report_date
        ]

        merged_records = merge_location_updates(
            existing_records,
            date_new_records,
            {report_date: update_scope[report_date]},
        )

        if cross_sheet:
            merged_records.sort(
                key=lambda record: (
                    record["location"].upper(),
                    record["area"].upper(),
                )
            )
        else:
            merged_records.sort(
                key=workbook_record_sort_key
            )

        if date_rows:
            start_row = min(date_rows)
            capacity_end = max(date_rows)
        else:
            start_row = find_location_insertion_row(
                worksheet,
                report_date,
                table_last_row,
            )
            capacity_end = start_row - 1

        needed_rows = len(merged_records)
        capacity = max(0, capacity_end - start_row + 1)

        # Reuse existing blank template rows before inserting anything.
        while capacity < needed_rows:
            next_row = start_row + capacity

            if next_row <= table_last_row and (
                next_row in date_rows
                or is_reusable_blank_location_row(
                    worksheet,
                    next_row,
                )
            ):
                capacity += 1
                capacity_end = next_row
                continue

            # A later populated row blocks the date block, so insert rows.
            insert_count = needed_rows - capacity
            worksheet.insert_rows(
                next_row,
                amount=insert_count,
            )
            table_last_row += insert_count
            capacity += insert_count
            capacity_end += insert_count
            break

        for offset, record in enumerate(merged_records):
            write_location_record(
                worksheet,
                start_row + offset,
                record,
                exact_style_row,
                uncertain_style_row,
            )

        # Clear unused placeholder/date rows but do not rewrite other dates.
        for row_number in range(
            start_row + needed_rows,
            capacity_end + 1,
        ):
            clear_location_row_values(
                worksheet,
                row_number,
            )

    resize_table(
        worksheet,
        table_name,
        f"A1:I{max(2, table_last_row)}",
    )

    return sum(
        1
        for row_number in range(2, table_last_row + 1)
        if parse_date_value(
            worksheet.cell(row_number, 1).value
        )
        is not None
        and clean_cell(
            worksheet.cell(row_number, 2).value
        )
        and clean_cell(
            worksheet.cell(row_number, 3).value
        )
    )


def read_daily_master_rows(
    worksheet: Any,
) -> dict[date, int]:
    """Map existing Daily Master dates to row numbers."""
    result: dict[date, int] = {}

    for row_number in range(2, worksheet.max_row + 1):
        report_date = parse_date_value(
            worksheet.cell(row_number, 1).value
        )
        if report_date is not None:
            result[report_date] = row_number

    return result


def find_first_blank_daily_row(
    worksheet: Any,
) -> int:
    """Find the first preformatted blank row in Daily Master."""
    for row_number in range(2, worksheet.max_row + 1):
        if parse_date_value(
            worksheet.cell(row_number, 1).value
        ) is None:
            return row_number

    return worksheet.max_row + 1


def update_daily_master(
    worksheet: Any,
    updates: dict[date, dict[str, int]],
) -> list[dict[str, Any]]:
    """Update only uploaded trades and leave other trade cells unchanged."""
    existing_rows = read_daily_master_rows(worksheet)
    sample_row = (
        min(existing_rows.values())
        if existing_rows
        else 2
    )
    changes: list[dict[str, Any]] = []

    for report_date in sorted(updates):
        row_number = existing_rows.get(report_date)

        if row_number is None:
            row_number = find_first_blank_daily_row(
                worksheet
            )
            copy_row_style(
                worksheet,
                sample_row,
                row_number,
                6,
            )
            worksheet.cell(
                row_number,
                1,
            ).value = report_date
            worksheet.cell(
                row_number,
                1,
            ).number_format = "yyyy-mm-dd"
            existing_rows[report_date] = row_number

        for trade, new_value in updates[report_date].items():
            column_number = DAILY_TRADE_COLUMN_MAP[trade]
            old_value = to_non_negative_int(
                worksheet.cell(
                    row_number,
                    column_number,
                ).value
            )

            worksheet.cell(
                row_number,
                column_number,
            ).value = new_value

            if old_value is None:
                action = "填入空白"
            elif old_value == new_value:
                action = "數值不變"
            else:
                action = "更新原有數值"

            changes.append(
                {
                    "日期": report_date.isoformat(),
                    "工種": trade,
                    "原有數值": old_value,
                    "新數值": new_value,
                    "操作": action,
                }
            )

        worksheet.cell(
            row_number,
            6,
        ).value = f"=SUM(B{row_number}:E{row_number})"

    last_used_row = max(existing_rows.values()) if existing_rows else 2

    resize_table(
        worksheet,
        "DailyMasterTable",
        f"A1:F{max(2, last_used_row)}",
    )

    return changes


def build_excel_location_records(
    summary_records: list[dict[str, Any]],
    *,
    review: bool,
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
]:
    """Convert app summaries into Location Detail and Cross-F records."""
    location_records: list[dict[str, Any]] = []
    cross_records: list[dict[str, Any]] = []

    for summary in summary_records:
        report_date = parse_date_value(
            summary.get("日期")
        )
        standard_location = clean_cell(
            summary.get("標準位置")
        )

        if report_date is None or not standard_location:
            continue

        trade_values = {
            trade: (
                to_non_negative_int(
                    summary.get(trade)
                )
                or 0
            )
            for trade in REQUIRED_TRADES
        }

        if sum(trade_values.values()) <= 0:
            continue

        if review and not is_single_area_uncertain_location(
            standard_location
        ):
            area, location, uncertain, note = (
                format_cross_floor_location(
                    standard_location
                )
            )

            cross_records.append(
                {
                    "date": report_date,
                    "area": area,
                    "location": location,
                    **trade_values,
                    "note": note,
                    "uncertain": uncertain,
                }
            )
            continue

        area, location, note, uncertain = (
            standard_location_to_workbook(
                standard_location
            )
        )

        location_records.append(
            {
                "date": report_date,
                "area": area,
                "location": location,
                **trade_values,
                "note": note,
                "uncertain": uncertain,
            }
        )

    return location_records, cross_records


def update_summary_formulas(
    workbook: Any,
) -> None:
    """Extend summary formulas to the current Daily Master table."""
    daily_sheet = workbook["Daily Master"]
    last_daily_row = 2

    for row_number in range(
        daily_sheet.max_row,
        1,
        -1,
    ):
        if parse_date_value(
            daily_sheet.cell(row_number, 1).value
        ) is not None:
            last_daily_row = row_number
            break

    department_sheet = workbook["Department Summary"]
    department_sheet["B2"] = (
        f"=SUM('Daily Master'!B2:B{last_daily_row})"
    )
    department_sheet["B3"] = (
        f"=SUM('Daily Master'!C2:C{last_daily_row})"
    )
    department_sheet["B4"] = (
        f"=SUM('Daily Master'!D2:D{last_daily_row})"
    )
    department_sheet["B5"] = (
        f"=SUM('Daily Master'!E2:E{last_daily_row})"
    )
    department_sheet["B6"] = "=SUM(B2:B5)"

    overview_sheet = workbook["Overview"]
    overview_sheet["B6"] = "=MAX('Daily Master'!A:A)"
    overview_sheet["B7"] = "='Department Summary'!B6"
    overview_sheet["B8"] = "='Department Summary'!B2"
    overview_sheet["B9"] = "='Department Summary'!B3"
    overview_sheet["B10"] = "='Department Summary'!B4"
    overview_sheet["B11"] = "='Department Summary'!B5"
    overview_sheet["B12"] = (
        f"=COUNTA('Daily Master'!A2:A{last_daily_row})"
    )


def export_updated_workbook(
    *,
    template_bytes: bytes,
    summary_rows: list[dict[str, Any]],
    confirmed_records: list[dict[str, Any]],
    review_records: list[dict[str, Any]],
) -> tuple[bytes, dict[str, Any]]:
    """Update a copy of the template and return downloadable XLSX bytes."""
    workbook = load_workbook(
        io.BytesIO(template_bytes),
        data_only=False,
    )

    missing_sheets = [
        sheet_name
        for sheet_name in EXCEL_REQUIRED_SHEETS
        if sheet_name not in workbook.sheetnames
    ]
    if missing_sheets:
        raise ValueError(
            "Excel模板缺少工作表："
            + ", ".join(missing_sheets)
        )

    daily_updates, update_scope, duplicate_keys = (
        build_daily_updates(summary_rows)
    )

    if not update_scope:
        raise ValueError(
            "分析結果中沒有可用的日期及工種，"
            "無法更新Excel。"
        )

    daily_changes = update_daily_master(
        workbook["Daily Master"],
        daily_updates,
    )

    confirmed_location_records, confirmed_cross_records = (
        build_excel_location_records(
            confirmed_records,
            review=False,
        )
    )
    review_location_records, review_cross_records = (
        build_excel_location_records(
            review_records,
            review=True,
        )
    )

    new_location_records = (
        confirmed_location_records
        + review_location_records
    )
    new_cross_records = (
        confirmed_cross_records
        + review_cross_records
    )

    location_sheet = workbook["Location Detail"]
    cross_sheet = workbook[
        "Cross-F & distribution U"
    ]

    location_row_count = update_location_sheet_preserving_history(
        location_sheet,
        new_location_records,
        update_scope,
        table_name="LocationDetailTable",
        cross_sheet=False,
    )
    cross_row_count = update_location_sheet_preserving_history(
        cross_sheet,
        new_cross_records,
        update_scope,
        table_name="CrossFloorUnspecifiedTable",
        cross_sheet=True,
    )

    update_summary_formulas(workbook)

    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except Exception:
        pass

    output_buffer = io.BytesIO()
    workbook.save(output_buffer)

    preview = {
        "daily_changes": daily_changes,
        "location_rows": location_row_count,
        "cross_rows": cross_row_count,
        "updated_dates": [
            report_date.isoformat()
            for report_date in sorted(update_scope)
        ],
        "updated_scope": [
            (
                report_date.isoformat()
                + "："
                + ", ".join(
                    sorted(update_scope[report_date])
                )
            )
            for report_date in sorted(update_scope)
        ],
        "duplicate_keys": duplicate_keys,
    }

    return output_buffer.getvalue(), preview






# =========================================================
# Manual location-review functions
# =========================================================


REVIEW_ACTIONS = (
    "尚未處理",
    "確認為單一位置",
    "確認保留原始分布",
)


def numbered_floor_value(value: str) -> int:
    match = re.search(r"(\d+)", clean_cell(value))
    return int(match.group(1)) if match else -1


def build_valid_location_options(
    config: dict[str, Any],
) -> list[str]:
    """Build exact-location choices using the project configuration."""
    options: list[str] = [""]

    tower_floors = sorted(
        config.get("tower_floors", []),
        key=numbered_floor_value,
        reverse=True,
    )
    for tower in config.get("towers", []):
        for floor in tower_floors:
            options.append(f"{tower} / {floor}")

    podium_floors = list(
        config.get("podium_floors", [])
    )
    podium_floors.sort(
        key=lambda floor: (
            floor.upper() == "GF",
            -numbered_floor_value(floor),
        )
    )
    for floor in podium_floors:
        options.append(f"Podium / {floor}")

    basement_floors = sorted(
        config.get("basement_floors", []),
        key=numbered_floor_value,
    )
    for floor in basement_floors:
        options.append(f"Basement / {floor}")

    for location in config.get(
        "special_locations",
        [],
    ):
        if location and location not in options:
            options.append(location)

    return options


def derive_exact_location_fields(
    standard_location: str,
) -> tuple[str, str]:
    """Derive Tower and Floor fields from a confirmed exact location."""
    location = clean_cell(standard_location)

    tower_match = re.fullmatch(
        r"(T\d+)\s*/\s*(\d+F)",
        location,
        re.I,
    )
    if tower_match:
        return (
            tower_match.group(1).upper(),
            tower_match.group(2).upper(),
        )

    podium_match = re.fullmatch(
        r"Podium\s*/\s*(GF|\d+F)",
        location,
        re.I,
    )
    if podium_match:
        return (
            "",
            podium_match.group(1).upper(),
        )

    basement_match = re.fullmatch(
        r"Basement\s*/\s*(B\d+)",
        location,
        re.I,
    )
    if basement_match:
        return (
            "",
            basement_match.group(1).upper(),
        )

    return "", location


def apply_manual_review_overrides(
    detail_df: pd.DataFrame,
    overrides: dict[str, dict[str, Any]],
) -> pd.DataFrame:
    """Apply saved human-review decisions to the editable detail table."""
    reviewed_df = detail_df.copy()

    if "明細ID" not in reviewed_df.columns:
        return reviewed_df

    for row_index, row in reviewed_df.iterrows():
        record_id = clean_cell(row.get("明細ID"))
        override = overrides.get(record_id)

        if not override:
            continue

        for column_name in (
            "標準位置",
            "樓座",
            "樓層",
            "位置狀態",
            "人工備註",
        ):
            if column_name in override:
                reviewed_df.at[
                    row_index,
                    column_name,
                ] = override[column_name]

    return reviewed_df


# =========================================================
# Validation and reconciliation functions
# =========================================================


def audit_number(value: Any) -> int | None:
    """Convert a report value into an integer for validation."""
    if value is None or value == "":
        return None

    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None

    if pd.isna(numeric):
        return None

    return int(round(numeric))


def build_validation_report(
    summary_rows: list[dict[str, Any]],
    edited_detail_df: pd.DataFrame,
) -> tuple[pd.DataFrame, int, int]:
    """Validate every analysed report against its edited location details.

    Hard blockers:
    - analysis failed;
    - date/trade/Worker missing;
    - location-detail total does not equal Worker;
    - Today Total Manpower is lower than Worker;
    - more than one different report exists for the same date and trade.

    Warnings:
    - some manpower remains in Cross-floor, Distribution U or
      unspecified locations.
    """
    audit_columns = [
        "狀態",
        "日期",
        "工種",
        "文件",
        "Today Total",
        "Worker",
        "明細合計",
        "差額",
        "待確認人數",
        "待確認列數",
        "已確認保留人數",
        "已確認保留列數",
        "核對說明",
    ]

    if not summary_rows:
        return (
            pd.DataFrame(columns=audit_columns),
            1,
            0,
        )

    working_df = edited_detail_df.copy()

    required_detail_columns = {
        "文件": "",
        "工種": "",
        "人數": 0,
        "位置狀態": "需人工確認位置",
    }
    for column_name, default_value in required_detail_columns.items():
        if column_name not in working_df.columns:
            working_df[column_name] = default_value

    working_df["文件"] = (
        working_df["文件"]
        .fillna("")
        .astype(str)
        .str.strip()
    )
    working_df["工種"] = (
        working_df["工種"]
        .fillna("")
        .astype(str)
        .str.upper()
        .str.strip()
    )
    working_df["人數"] = pd.to_numeric(
        working_df["人數"],
        errors="coerce",
    ).fillna(0)
    working_df["位置狀態"] = (
        working_df["位置狀態"]
        .fillna("需人工確認位置")
        .astype(str)
        .str.strip()
    )

    date_trade_counts: dict[tuple[str, str], int] = {}
    for summary in summary_rows:
        date_text = clean_cell(summary.get("日期"))
        trade = clean_cell(summary.get("工種")).upper()
        if date_text and trade in REQUIRED_TRADES:
            key = (date_text, trade)
            date_trade_counts[key] = (
                date_trade_counts.get(key, 0) + 1
            )

    audit_rows: list[dict[str, Any]] = []
    blocker_count = 0
    warning_count = 0

    for summary in summary_rows:
        filename = clean_cell(summary.get("文件"))
        date_text = clean_cell(summary.get("日期"))
        trade = clean_cell(summary.get("工種")).upper()
        today_total = audit_number(
            summary.get("Today Total Manpower")
        )
        worker = audit_number(summary.get("Worker"))

        file_detail_df = working_df[
            working_df["文件"] == filename
        ].copy()

        detail_total = int(
            file_detail_df["人數"].sum()
        ) if not file_detail_df.empty else 0

        retained_mask = (
            file_detail_df["位置狀態"]
            == "已人工確認保留"
        )
        review_mask = ~file_detail_df[
            "位置狀態"
        ].isin(
            {
                "已解析",
                "已人工確認保留",
            }
        )

        review_people = int(
            file_detail_df.loc[
                review_mask,
                "人數",
            ].sum()
        ) if not file_detail_df.empty else 0
        review_rows = int(review_mask.sum())

        retained_people = int(
            file_detail_df.loc[
                retained_mask,
                "人數",
            ].sum()
        ) if not file_detail_df.empty else 0
        retained_rows = int(retained_mask.sum())

        reasons: list[str] = []
        is_blocker = False

        reconciliation_text = clean_cell(
            summary.get("Worker核對")
        )
        extraction_method = clean_cell(
            summary.get("明細提取方式")
        )

        if (
            date_text in {"", "未讀取", "分析失敗"}
            or extraction_method == "失敗"
            or reconciliation_text.startswith("錯誤")
        ):
            reasons.append("報告分析失敗或日期未能讀取")
            is_blocker = True

        if trade not in REQUIRED_TRADES:
            reasons.append("工種未能識別")
            is_blocker = True

        if worker is None:
            reasons.append("未能讀取Worker")
            is_blocker = True
            difference: int | None = None
        else:
            difference = detail_total - worker
            if difference != 0:
                reasons.append(
                    f"明細與Worker相差{difference:+d}"
                )
                is_blocker = True

        if (
            today_total is not None
            and worker is not None
            and today_total < worker
        ):
            reasons.append(
                "Today Total Manpower低於Worker"
            )
            is_blocker = True

        if date_trade_counts.get(
            (date_text, trade),
            0,
        ) > 1:
            reasons.append("同一日期及工種有多份不同報告")
            is_blocker = True

        if is_blocker:
            status = "❌ 阻止匯出"
            blocker_count += 1
        elif review_people > 0:
            status = "⚠️ 可匯出，仍待覆核"
            warning_count += 1
            reasons.append(
                "人數已核對，但部分位置尚未完成人工覆核"
            )
        elif retained_people > 0:
            status = "✅ 通過，已確認保留"
            reasons.append(
                "人工確認保留Cross-floor／Distribution U／"
                "未指定位置，不作假設分配"
            )
        else:
            status = "✅ 通過"
            reasons.append("Worker及位置明細核對通過")

        audit_rows.append(
            {
                "狀態": status,
                "日期": date_text,
                "工種": trade,
                "文件": filename,
                "Today Total": today_total,
                "Worker": worker,
                "明細合計": detail_total,
                "差額": difference,
                "待確認人數": review_people,
                "待確認列數": review_rows,
                "已確認保留人數": retained_people,
                "已確認保留列數": retained_rows,
                "核對說明": "；".join(reasons),
            }
        )

    audit_df = pd.DataFrame(
        audit_rows,
        columns=audit_columns,
    )

    return audit_df, blocker_count, warning_count


# =========================================================
# Page title
# =========================================================

st.title("🏗️ 工地人力日報智能體（Python版）")
st.write(
    "建立工程及樓層配置，上傳 AC、EL、FS 和 PD 日報。"
    "目前以純 Python 分析 AC、EL、FS 和 PD PDF，不需要任何 AI API。"
    "系統可自動偵測姓名表、Manpower數字欄、描述中的人數，或只保留總人數。"
)
st.divider()


# =========================================================
# Section 1: Project configuration
# =========================================================

st.subheader("1. 工程項目設定")

with st.form("project_configuration_form"):
    project_name = st.text_input(
        "工程名稱",
        placeholder="例如：土瓜灣鴻福街及銀漢街發展項目",
    )
    tower_input = st.text_input(
        "樓座名稱",
        value="T1, T2, T3",
        help="用逗號分隔，例如：T1, T2, T3",
    )

    col1, col2 = st.columns(2)

    with col1:
        basement_input = st.text_input(
            "Basement 樓層",
            value="B4, B3, B2, B1",
        )
        podium_input = st.text_input(
            "Podium 樓層",
            value="GF, 1F, 2F, 3F",
        )
        special_floor_input = st.text_input(
            "特殊樓層或區域",
            value="Roof",
        )

    with col2:
        tower_start_floor = st.number_input(
            "Tower 最低樓層編號",
            min_value=1,
            max_value=200,
            value=5,
            step=1,
        )
        tower_end_floor = st.number_input(
            "Tower 最高樓層編號",
            min_value=1,
            max_value=200,
            value=29,
            step=1,
        )
        excluded_tower_floor_input = st.text_input(
            "不存在的 Tower 樓層",
            value="13F, 14F, 24F",
        )

    merge_podium_towers = st.checkbox(
        "Podium 樓層忽略 Tower 編號並合併",
        value=True,
    )

    submitted = st.form_submit_button(
        "儲存工程設定",
        type="primary",
        width="stretch",
    )

if submitted:
    if not project_name.strip():
        st.error("請輸入工程名稱。")
    elif tower_end_floor < tower_start_floor:
        st.error("Tower 最高樓層不能低於最低樓層。")
    else:
        towers = [tower.upper() for tower in text_to_list(tower_input)]
        basement_floors = [
            normalize_floor(item) for item in text_to_list(basement_input)
        ]
        podium_floors = [
            normalize_floor(item) for item in text_to_list(podium_input)
        ]
        special_locations = [
            normalize_floor(item) for item in text_to_list(special_floor_input)
        ]
        excluded_tower_floors = {
            normalize_floor(item)
            for item in text_to_list(excluded_tower_floor_input)
        }
        tower_floors = [
            f"{floor_number}F"
            for floor_number in range(
                int(tower_start_floor),
                int(tower_end_floor) + 1,
            )
            if f"{floor_number}F" not in excluded_tower_floors
        ]

        st.session_state["project_config"] = {
            "project_name": project_name.strip(),
            "towers": towers,
            "basement_floors": basement_floors,
            "podium_floors": podium_floors,
            "tower_start_floor": int(tower_start_floor),
            "tower_end_floor": int(tower_end_floor),
            "excluded_tower_floors": sorted(
                excluded_tower_floors,
                key=floor_sort_key,
            ),
            "tower_floors": tower_floors,
            "special_locations": special_locations,
            "merge_podium_tower_references": merge_podium_towers,
        }
        st.session_state.pop("analysis_results", None)
        st.success("工程設定已儲存。")


# =========================================================
# Section 2: Saved configuration
# =========================================================

if "project_config" in st.session_state:
    st.divider()
    st.subheader("2. 已儲存的工程設定")
    config = st.session_state["project_config"]

    col1, col2, col3 = st.columns(3)

    with col1:
        st.write(f"**工程名稱：** {config['project_name']}")
        st.write("**樓座：** " + (", ".join(config["towers"]) or "未設定"))

    with col2:
        st.write(
            "**Basement：** "
            + (", ".join(config["basement_floors"]) or "未設定")
        )
        st.write(
            "**Podium：** "
            + (", ".join(config["podium_floors"]) or "未設定")
        )
        st.write(
            "**特殊位置：** "
            + (", ".join(config["special_locations"]) or "未設定")
        )

    with col3:
        st.write(
            f"**Tower 範圍：** {config['tower_start_floor']}F–"
            f"{config['tower_end_floor']}F"
        )
        st.write(
            "**不存在樓層：** "
            + (", ".join(config["excluded_tower_floors"]) or "沒有")
        )
        st.write(f"**有效 Tower 樓層：** {len(config['tower_floors'])} 層")

    config_json = json.dumps(config, ensure_ascii=False, indent=2)
    config_download_col, template_download_col = (
        st.columns(2)
    )

    with config_download_col:
        st.download_button(
            "下載工程設定 JSON",
            data=config_json,
            file_name="project_config.json",
            mime="application/json",
            width="stretch",
        )

    with template_download_col:
        blank_template_bytes = (
            create_blank_master_template_bytes(
                config
            )
        )
        blank_project_filename = (
            safe_filename_part(
                config["project_name"]
            )
            + "_Manpower_Blank_Template.xlsx"
        )
        st.download_button(
            "下載全新空白Excel模板",
            data=blank_template_bytes,
            file_name=blank_project_filename,
            mime=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
            width="stretch",
        )

    st.caption(
        "空白模板保留六張工作表、深藍表頭、公式、"
        "表格格式及黃色不確定位置規則，但不包含舊工程歷史資料。"
    )


# =========================================================
# Section 3: Report upload
# =========================================================

st.divider()
st.subheader("3. 上傳工地日報")

project_is_ready = "project_config" in st.session_state
if not project_is_ready:
    st.warning("請先完成並儲存工程設定，再上傳日報。")

uploaded_files = st.file_uploader(
    "選擇 AC、EL、FS、PD 日報",
    type=["pdf", "png", "jpg", "jpeg", "xlsx", "xls"],
    accept_multiple_files=True,
    disabled=not project_is_ready,
)


# =========================================================
# Section 4: File checking
# =========================================================

if uploaded_files:
    st.divider()
    st.subheader("4. 文件檢查結果")

    seen_hashes: dict[str, str] = {}
    file_records: list[dict[str, Any]] = []
    unique_files: list[dict[str, Any]] = []
    duplicate_files: list[dict[str, Any]] = []

    for uploaded_file in uploaded_files:
        file_bytes = uploaded_file.getvalue()
        file_hash = calculate_file_hash(file_bytes)
        detected_trade = detect_trade_from_filename(uploaded_file.name)

        if file_hash in seen_hashes:
            original_filename = seen_hashes[file_hash]
            file_records.append(
                {
                    "文件名稱": uploaded_file.name,
                    "工種": detected_trade or "未識別",
                    "大小": format_file_size(len(file_bytes)),
                    "狀態": "重複，已忽略",
                    "備註": f"與 {original_filename} 完全相同",
                }
            )
            duplicate_files.append(
                {
                    "name": uploaded_file.name,
                    "original_name": original_filename,
                }
            )
        else:
            seen_hashes[file_hash] = uploaded_file.name
            file_records.append(
                {
                    "文件名稱": uploaded_file.name,
                    "工種": detected_trade or "未識別",
                    "大小": format_file_size(len(file_bytes)),
                    "狀態": "保留",
                    "備註": "",
                }
            )
            unique_files.append(
                {
                    "name": uploaded_file.name,
                    "bytes": file_bytes,
                    "hash": file_hash,
                    "trade": detected_trade,
                    "size": len(file_bytes),
                    "type": uploaded_file.type,
                }
            )

    st.session_state["unique_uploaded_files"] = unique_files

    metric1, metric2, metric3 = st.columns(3)
    metric1.metric("上傳文件", len(uploaded_files))
    metric2.metric("有效文件", len(unique_files))
    metric3.metric("重複文件", len(duplicate_files))

    st.dataframe(file_records, width="stretch", hide_index=True)

    if duplicate_files:
        for duplicate in duplicate_files:
            st.warning(
                f"{duplicate['name']} 與 {duplicate['original_name']} 完全相同，"
                "第二份已忽略。"
            )
    else:
        st.success("沒有發現完全重複的文件。")

    detected_trades = {
        file_data["trade"]
        for file_data in unique_files
        if file_data["trade"] is not None
    }
    missing_trades = set(REQUIRED_TRADES) - detected_trades

    trade_columns = st.columns(4)
    for column, trade in zip(trade_columns, REQUIRED_TRADES):
        with column:
            if trade in detected_trades:
                st.success(f"{trade}：已上傳")
            else:
                st.error(f"{trade}：缺少")

    if missing_trades:
        st.warning("目前缺少以下工種報告：" + ", ".join(sorted(missing_trades)))
    else:
        st.success("AC、EL、FS 和 PD 四個工種的報告均已上傳。")

    st.markdown("#### 準備分析的有效文件")
    for index, file_data in enumerate(unique_files, start=1):
        st.write(
            f"{index}. {file_data['name']} — "
            f"{file_data['trade'] or '未識別'} — "
            f"{format_file_size(file_data['size'])}"
        )
else:
    st.info("尚未上傳日報。請選擇 PDF、圖片或 Excel 文件。")


# =========================================================
# Section 5: Flexible Python report analysis
# =========================================================

st.divider()
st.subheader("5. AC／EL／FS／PD 彈性 Python 分析")
st.caption(f"目前版本：{APP_VERSION}")

unique_uploaded_files = st.session_state.get("unique_uploaded_files", [])
supported_pdf_files = [
    file_data
    for file_data in unique_uploaded_files
    if file_data.get("trade") in REQUIRED_TRADES
    and file_data["name"].lower().endswith(".pdf")
]

if not project_is_ready:
    st.warning("請先儲存工程設定。")
elif not supported_pdf_files:
    st.info("請先上傳至少一份 AC、EL、FS 或 PD PDF 日報。")
else:
    st.write(
        "準備分析："
        + ", ".join(file_data["name"] for file_data in supported_pdf_files)
    )

    st.markdown("### 各工種人數提取方式")
    st.caption(
        "建議保持『自動偵測』。其他工程沒有工人姓名表時，"
        "系統會改找Location＋Manpower數字欄、描述中的X人，"
        "再不成功才只保留總人數並要求人工分配位置。"
    )

    uploaded_trades = [
        trade
        for trade in REQUIRED_TRADES
        if any(file_data.get("trade") == trade for file_data in supported_pdf_files)
    ]

    parser_modes: dict[str, str] = {}
    parser_columns = st.columns(max(1, len(uploaded_trades)))
    option_labels = list(PARSER_MODE_OPTIONS.keys())

    for column, trade in zip(parser_columns, uploaded_trades):
        with column:
            selected_label = st.selectbox(
                f"{trade} 提取方式",
                options=option_labels,
                index=0,
                key=f"parser_mode_{trade}",
            )
            parser_modes[trade] = PARSER_MODE_OPTIONS[selected_label]

    if st.button(
        "使用 Python 分析全部 PDF",
        type="primary",
        width="stretch",
    ):
        results: list[dict[str, Any]] = []

        with st.spinner("正在辨認報告格式、提取位置及核對人數..."):
            for file_data in supported_pdf_files:
                trade = file_data.get("trade") or ""

                try:
                    result = analyse_report_pdf(
                        file_data,
                        st.session_state["project_config"],
                        trade,
                        parser_modes.get(trade, "auto"),
                    )
                    results.append(result)

                except Exception as error:
                    results.append(
                        {
                            "summary": {
                                "文件": file_data["name"],
                                "日期": "分析失敗",
                                "工種": trade or "未識別",
                                "Ref. No.": "",
                                "Today Total Manpower": None,
                                "Worker": None,
                                "管理及技術人員": None,
                                "工作明細人數合計": None,
                                "Worker核對": f"錯誤：{error}",
                                "明細提取方式": "失敗",
                                "明細列數": 0,
                                "重複姓名": "未檢查",
                            },
                            "details": [],
                            "raw_text": "",
                        }
                    )

        st.session_state["analysis_results"] = results
        st.session_state.pop(
            "manual_review_overrides_v061",
            None,
        )
        st.session_state.pop(
            "validation_records_v061",
            None,
        )
        st.session_state.pop(
            "excel_export_bytes_v070",
            None,
        )
        st.session_state.pop(
            "excel_export_preview_v070",
            None,
        )
        st.success("AC／EL／FS／PD PDF 分析完成。")


analysis_results = st.session_state.get("analysis_results", [])

if analysis_results:
    st.markdown("### 分析摘要")
    summary_rows = [result["summary"] for result in analysis_results]
    summary_df = pd.DataFrame(summary_rows)
    st.dataframe(summary_df, width="stretch", hide_index=True)

    for summary in summary_rows:
        reconciliation = summary.get("Worker核對", "")
        if reconciliation == "一致":
            st.success(f"{summary['文件']}：明細合計與 Worker 一致。")
        elif reconciliation == "位置未分配":
            st.warning(
                f"{summary['文件']}：只取得總人數，位置需要人工分配。"
            )
        else:
            st.warning(f"{summary['文件']}：{reconciliation}")

        duplicate_names = summary.get("重複姓名", "沒有")
        if duplicate_names not in {"沒有", "不適用", "未檢查"}:
            st.warning(
                f"{summary['文件']}：Worker 表出現重複姓名：{duplicate_names}"
            )

    all_detail_rows = [
        row
        for result in analysis_results
        for row in result["details"]
    ]

    st.markdown("### 工作位置及人數明細（可人工修改）")

    if all_detail_rows:
        detail_df = pd.DataFrame(all_detail_rows)
        detail_df.insert(
            0,
            "明細ID",
            [
                f"R{index:04d}"
                for index in range(
                    1,
                    len(detail_df) + 1,
                )
            ],
        )
        if "人工備註" not in detail_df.columns:
            detail_df["人工備註"] = ""

        preferred_columns = [
            "明細ID",
            "日期",
            "工種",
            "文件",
            "Item",
            "工作描述",
            "工人姓名",
            "原始位置",
            "樓座",
            "樓層",
            "標準位置",
            "人數",
            "人數計算方法",
            "位置狀態",
            "人工備註",
        ]
        detail_df = detail_df.reindex(columns=preferred_columns)

        edited_detail_df = st.data_editor(
            detail_df,
            width="stretch",
            hide_index=True,
            num_rows="dynamic",
            disabled=[
                "明細ID",
                "日期",
                "工種",
                "文件",
                "Item",
                "工作描述",
                "工人姓名",
                "原始位置",
                "樓座",
                "樓層",
                "人數計算方法",
            ],
            column_config={
                "明細ID": None,
                "人數": st.column_config.NumberColumn(
                    "人數",
                    min_value=0,
                    step=1,
                )
            },
            key="combined_detail_editor_v061",
        )

        # Fill IDs for any manually added rows.
        for row_index in edited_detail_df.index:
            if not clean_cell(
                edited_detail_df.at[
                    row_index,
                    "明細ID",
                ]
            ):
                edited_detail_df.at[
                    row_index,
                    "明細ID",
                ] = f"M{int(row_index) + 1:04d}"

        saved_overrides = st.session_state.get(
            "manual_review_overrides_v061",
            {},
        )
        edited_detail_df = (
            apply_manual_review_overrides(
                edited_detail_df,
                saved_overrides,
            )
        )

        # =================================================
        # Section 6: Manual location review
        # =================================================

        st.divider()
        st.subheader("6. 待確認位置處理")

        unresolved_mask = ~edited_detail_df[
            "位置狀態"
        ].isin(
            {
                "已解析",
                "已人工確認保留",
            }
        )
        retained_mask = (
            edited_detail_df["位置狀態"]
            == "已人工確認保留"
        )

        review_metric1, review_metric2, review_metric3 = (
            st.columns(3)
        )
        review_metric1.metric(
            "尚未處理列數",
            int(unresolved_mask.sum()),
        )
        review_metric2.metric(
            "尚未處理人數",
            int(
                pd.to_numeric(
                    edited_detail_df.loc[
                        unresolved_mask,
                        "人數",
                    ],
                    errors="coerce",
                ).fillna(0).sum()
            ),
        )
        review_metric3.metric(
            "已確認保留列數",
            int(retained_mask.sum()),
        )

        reviewable_mask = (
            edited_detail_df["位置狀態"]
            != "已解析"
        )
        review_source_df = edited_detail_df.loc[
            reviewable_mask
        ].copy()

        if review_source_df.empty:
            st.success(
                "沒有Cross-floor、Distribution U或"
                "未指定位置需要處理。"
            )
        else:
            st.info(
                "知道實際單一位置時，選「確認為單一位置」；"
                "報告本身沒有提供分布時，選"
                "「確認保留原始分布」。後者仍會放入"
                "Cross-F／Distribution U，不會被系統亂分配。"
            )

            valid_location_options = (
                build_valid_location_options(
                    st.session_state[
                        "project_config"
                    ]
                )
            )

            review_rows: list[dict[str, Any]] = []
            for _, review_row in review_source_df.iterrows():
                record_id = clean_cell(
                    review_row.get("明細ID")
                )
                saved_override = saved_overrides.get(
                    record_id,
                    {},
                )
                current_status = clean_cell(
                    review_row.get("位置狀態")
                )

                if current_status == "已人工確認保留":
                    default_action = (
                        "確認保留原始分布"
                    )
                elif (
                    current_status == "已解析"
                    and clean_cell(
                        review_row.get("標準位置")
                    )
                    in valid_location_options
                ):
                    default_action = (
                        "確認為單一位置"
                    )
                else:
                    default_action = "尚未處理"

                review_rows.append(
                    {
                        "明細ID": record_id,
                        "日期": review_row.get("日期", ""),
                        "工種": review_row.get("工種", ""),
                        "文件": review_row.get("文件", ""),
                        "工作描述／工人": (
                            clean_cell(
                                review_row.get(
                                    "工作描述"
                                )
                            )
                            or clean_cell(
                                review_row.get(
                                    "工人姓名"
                                )
                            )
                        ),
                        "原始位置": review_row.get(
                            "原始位置",
                            "",
                        ),
                        "人數": review_row.get("人數", 0),
                        "目前標準位置": review_row.get(
                            "標準位置",
                            "",
                        ),
                        "目前狀態": current_status,
                        "處理方式": saved_override.get(
                            "處理方式",
                            default_action,
                        ),
                        "確認為位置": saved_override.get(
                            "確認為位置",
                            (
                                review_row.get(
                                    "標準位置",
                                    "",
                                )
                                if review_row.get(
                                    "標準位置",
                                    "",
                                )
                                in valid_location_options
                                else ""
                            ),
                        ),
                        "人工備註": saved_override.get(
                            "人工備註",
                            review_row.get(
                                "人工備註",
                                "",
                            ),
                        ),
                    }
                )

            review_editor_df = st.data_editor(
                pd.DataFrame(review_rows),
                width="stretch",
                hide_index=True,
                disabled=[
                    "明細ID",
                    "日期",
                    "工種",
                    "文件",
                    "工作描述／工人",
                    "原始位置",
                    "人數",
                    "目前標準位置",
                    "目前狀態",
                ],
                column_config={
                    "明細ID": None,
                    "處理方式": (
                        st.column_config.SelectboxColumn(
                            "處理方式",
                            options=list(REVIEW_ACTIONS),
                            required=True,
                        )
                    ),
                    "確認為位置": (
                        st.column_config.SelectboxColumn(
                            "確認為位置",
                            options=valid_location_options,
                        )
                    ),
                },
                key="manual_review_editor_v061",
            )

            apply_col, clear_col = st.columns(2)

            with apply_col:
                apply_review = st.button(
                    "套用人工確認",
                    type="primary",
                    width="stretch",
                )

            with clear_col:
                clear_review = st.button(
                    "清除全部人工確認",
                    width="stretch",
                )

            if clear_review:
                st.session_state.pop(
                    "manual_review_overrides_v061",
                    None,
                )
                st.rerun()

            if apply_review:
                new_overrides = dict(saved_overrides)
                review_errors: list[str] = []

                for _, reviewed_row in (
                    review_editor_df.iterrows()
                ):
                    record_id = clean_cell(
                        reviewed_row.get("明細ID")
                    )
                    action = clean_cell(
                        reviewed_row.get("處理方式")
                    )
                    selected_location = clean_cell(
                        reviewed_row.get(
                            "確認為位置"
                        )
                    )
                    manual_note = clean_cell(
                        reviewed_row.get("人工備註")
                    )

                    if action == "尚未處理":
                        new_overrides.pop(
                            record_id,
                            None,
                        )
                        continue

                    if action == "確認為單一位置":
                        if not selected_location:
                            review_errors.append(
                                f"{record_id}：請選擇確認位置"
                            )
                            continue

                        tower_value, floor_value = (
                            derive_exact_location_fields(
                                selected_location
                            )
                        )
                        new_overrides[record_id] = {
                            "處理方式": action,
                            "確認為位置": (
                                selected_location
                            ),
                            "標準位置": (
                                selected_location
                            ),
                            "樓座": tower_value,
                            "樓層": floor_value,
                            "位置狀態": "已解析",
                            "人工備註": manual_note,
                        }
                        continue

                    if action == "確認保留原始分布":
                        new_overrides[record_id] = {
                            "處理方式": action,
                            "確認為位置": "",
                            "標準位置": clean_cell(
                                reviewed_row.get(
                                    "目前標準位置"
                                )
                            ),
                            "樓座": clean_cell(
                                review_source_df.loc[
                                    review_source_df[
                                        "明細ID"
                                    ]
                                    == record_id,
                                    "樓座",
                                ].iloc[0]
                            ),
                            "樓層": clean_cell(
                                review_source_df.loc[
                                    review_source_df[
                                        "明細ID"
                                    ]
                                    == record_id,
                                    "樓層",
                                ].iloc[0]
                            ),
                            "位置狀態": (
                                "已人工確認保留"
                            ),
                            "人工備註": manual_note,
                        }

                if review_errors:
                    st.error(
                        "未能套用："
                        + "；".join(review_errors)
                    )
                else:
                    st.session_state[
                        "manual_review_overrides_v061"
                    ] = new_overrides
                    st.success("人工確認已套用。")
                    st.rerun()

        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                "下載分析摘要 CSV",
                data=summary_df.to_csv(index=False).encode("utf-8-sig"),
                file_name="report_analysis_summary.csv",
                mime="text/csv",
                width="stretch",
            )
        with col2:
            st.download_button(
                "下載位置明細 CSV",
                data=edited_detail_df.to_csv(index=False).encode("utf-8-sig"),
                file_name="location_details.csv",
                mime="text/csv",
                width="stretch",
            )

        st.session_state["edited_detail_records"] = (
            edited_detail_df.to_dict("records")
        )

        # =================================================
        # Section 7: Validation centre
        # =================================================

        st.divider()
        st.subheader("7. 資料核對中心")

        (
            audit_df,
            audit_blocker_count,
            audit_warning_count,
        ) = build_validation_report(
            summary_rows,
            edited_detail_df,
        )

        st.session_state[
            "validation_records_v061"
        ] = audit_df.to_dict("records")
        st.session_state[
            "validation_blockers_v061"
        ] = audit_blocker_count
        st.session_state[
            "validation_warnings_v061"
        ] = audit_warning_count

        audit_metric1, audit_metric2, audit_metric3 = (
            st.columns(3)
        )
        audit_metric1.metric(
            "分析報告",
            len(audit_df),
        )
        audit_metric2.metric(
            "阻止匯出",
            audit_blocker_count,
        )
        audit_metric3.metric(
            "尚未完成覆核",
            audit_warning_count,
        )

        st.dataframe(
            audit_df,
            width="stretch",
            hide_index=True,
        )

        st.download_button(
            "下載資料核對報告 CSV",
            data=audit_df.to_csv(
                index=False
            ).encode("utf-8-sig"),
            file_name="manpower_validation_report.csv",
            mime="text/csv",
            width="stretch",
        )

        if audit_blocker_count > 0:
            st.error(
                "存在人數差額、分析失敗、缺少Worker，"
                "或同一日期工種重複。"
                "修正前Excel匯出會被停用。"
            )
        elif audit_warning_count > 0:
            st.warning(
                "所有Worker人數均已核對一致，"
                "但仍有位置尚未完成人工覆核。"
                "請到上方「待確認位置處理」選擇"
                "確認單一位置或確認保留原始分布。"
            )
        else:
            st.success(
                "所有報告的人數及位置明細均通過核對。"
            )

        # =================================================
        # Section 8: Merge same-date and same-location data
        # =================================================

        st.divider()
        st.subheader("8. 跨工種位置合併")

        confirmed_summary_df, review_summary_df = (
            build_location_summary(edited_detail_df)
        )

        st.session_state["confirmed_summary_records"] = (
            confirmed_summary_df.to_dict("records")
        )
        st.session_state["review_summary_records"] = (
            review_summary_df.to_dict("records")
        )

        if not confirmed_summary_df.empty:
            total_confirmed = int(confirmed_summary_df["合計"].sum())

            metric1, metric2 = st.columns(2)
            metric1.metric("已合併標準位置", len(confirmed_summary_df))
            metric2.metric("已解析位置人數合計", total_confirmed)

            st.markdown("### 已解析位置摘要")
            st.caption(
                "合併條件：同一日期＋同一標準位置。"
                "排序：T1、T2、T3依次排列，每座由高層至低層；"
                "Podium由高層至GF；Basement由B1至B2、B3、B4。"
            )
            st.dataframe(
                confirmed_summary_df,
                width="stretch",
                hide_index=True,
            )

            st.download_button(
                "下載跨工種合併摘要 CSV",
                data=confirmed_summary_df.to_csv(index=False).encode("utf-8-sig"),
                file_name="merged_trade_location_summary.csv",
                mime="text/csv",
                width="stretch",
            )
        else:
            st.warning("目前沒有可直接合併的已解析位置。")

        st.markdown("### Cross-floor／Distribution U／未指定位置")
        st.caption(
            "只有總數、只寫樓座、跨多個區域或無法確定樓層的資料會保留在這裡。"
            "系統不會擅自平均分配。"
        )

        if not review_summary_df.empty:
            st.dataframe(
                review_summary_df,
                width="stretch",
                hide_index=True,
            )

            st.download_button(
                "下載待確認位置摘要 CSV",
                data=review_summary_df.to_csv(index=False).encode("utf-8-sig"),
                file_name="cross_floor_distribution_u.csv",
                mime="text/csv",
                width="stretch",
            )
        else:
            st.success("沒有 Cross-floor、Distribution U 或未指定位置。")

    else:
        st.error("未能從 PDF 提取任何可計算明細。")

    with st.expander("查看 PDF 原始提取文字（除錯用）", expanded=False):
        for result in analysis_results:
            st.markdown(f"#### {result['summary']['文件']}")
            st.text(result["raw_text"] or "沒有提取到文字。")

# =========================================================
# Section 9: Excel workbook export
# =========================================================

st.divider()
st.subheader("9. Excel工作簿輸出")

if not analysis_results:
    st.info("請先完成PDF分析，才可以更新Excel。")
else:
    confirmed_export_records = st.session_state.get(
        "confirmed_summary_records",
        [],
    )
    review_export_records = st.session_state.get(
        "review_summary_records",
        [],
    )
    audit_blocker_count = st.session_state.get(
        "validation_blockers_v061",
        1,
    )
    audit_warning_count = st.session_state.get(
        "validation_warnings_v061",
        0,
    )

    if audit_blocker_count > 0:
        st.error(
            "資料核對中心仍有阻止匯出的錯誤。"
            "請先修正人數或報告問題。"
        )
    elif audit_warning_count > 0:
        st.warning(
            "仍有位置尚未完成人工覆核。"
            "雖然不阻止匯出，建議先在"
            "「待確認位置處理」完成決定。"
        )
    else:
        st.success("資料核對已通過，可以建立Excel。")

    export_mode = st.radio(
        "選擇Excel輸出模式",
        options=[
            "更新現有工程Excel",
            "建立全新工程Excel",
        ],
        horizontal=True,
        key="excel_export_mode_v070",
    )

    template_bytes_for_export: bytes | None = None
    template_name_for_export = ""

    if export_mode == "更新現有工程Excel":
        st.write(
            "上傳正在使用的Manpower Excel。"
            "系統只會產生一份新副本，不會改動原始檔。"
        )

        excel_template_file = st.file_uploader(
            "上傳現有Excel（.xlsx）",
            type=["xlsx"],
            accept_multiple_files=False,
            key="excel_template_uploader_v070",
        )

        if excel_template_file is not None:
            template_bytes_for_export = (
                excel_template_file.getvalue()
            )
            template_name_for_export = (
                excel_template_file.name
            )
            st.success(
                f"已讀取現有Excel："
                f"{excel_template_file.name}"
            )
            st.caption(
                "同一日期只更新本次已分析的工種；"
                "其他工種及較早日期保持原樣。"
            )

    else:
        project_config_for_excel = (
            st.session_state["project_config"]
        )
        template_bytes_for_export = (
            create_blank_master_template_bytes(
                project_config_for_excel
            )
        )
        template_name_for_export = (
            safe_filename_part(
                project_config_for_excel[
                    "project_name"
                ]
            )
            + "_Manpower.xlsx"
        )

        st.success(
            "將使用內置空白Master Template建立新工程Excel。"
        )
        st.caption(
            "新檔不包含舊工程歷史資料，"
            "但保留六張工作表、公式、深藍表頭、"
            "藍色間隔列及黃色不確定位置格式。"
        )

        st.download_button(
            "下載空白Master Template",
            data=template_bytes_for_export,
            file_name=(
                safe_filename_part(
                    project_config_for_excel[
                        "project_name"
                    ]
                )
                + "_Manpower_Blank_Template.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
            width="stretch",
        )

    preview_button_text = (
        "建立現有Excel更新預覽"
        if export_mode
        == "更新現有工程Excel"
        else "建立全新工程Excel預覽"
    )

    if template_bytes_for_export is not None:
        if st.button(
            preview_button_text,
            type="primary",
            width="stretch",
            disabled=(audit_blocker_count > 0),
        ):
            try:
                export_bytes, export_preview = (
                    export_updated_workbook(
                        template_bytes=(
                            template_bytes_for_export
                        ),
                        summary_rows=summary_rows,
                        confirmed_records=(
                            confirmed_export_records
                        ),
                        review_records=review_export_records,
                    )
                )

                st.session_state[
                    "excel_export_bytes_v070"
                ] = export_bytes
                st.session_state[
                    "excel_export_preview_v070"
                ] = export_preview
                st.session_state[
                    "excel_export_template_name_v070"
                ] = template_name_for_export
                st.session_state[
                    "excel_export_mode_saved_v070"
                ] = export_mode

                st.success(
                    "Excel預覽已建立。"
                )

            except Exception as error:
                st.session_state.pop(
                    "excel_export_bytes_v070",
                    None,
                )
                st.session_state.pop(
                    "excel_export_preview_v070",
                    None,
                )
                st.error("建立Excel時發生錯誤。")
                st.exception(error)

    export_bytes = st.session_state.get(
        "excel_export_bytes_v070"
    )
    export_preview = st.session_state.get(
        "excel_export_preview_v070"
    )

    if export_bytes and export_preview:
        st.markdown("### 預計更新內容")

        metric1, metric2, metric3 = st.columns(3)
        metric1.metric(
            "更新日期數量",
            len(export_preview["updated_dates"]),
        )
        metric2.metric(
            "Location Detail總列數",
            export_preview["location_rows"],
        )
        metric3.metric(
            "Cross-F總列數",
            export_preview["cross_rows"],
        )

        for scope_text in export_preview[
            "updated_scope"
        ]:
            st.info(scope_text)

        daily_change_df = pd.DataFrame(
            export_preview["daily_changes"]
        )

        if not daily_change_df.empty:
            st.markdown("#### Daily Master變更預覽")
            st.dataframe(
                daily_change_df,
                width="stretch",
                hide_index=True,
            )

        duplicate_keys = export_preview.get(
            "duplicate_keys",
            [],
        )
        if duplicate_keys:
            st.warning(
                "同一日期及工種出現多份分析結果，"
                "目前使用最後一份："
                + ", ".join(duplicate_keys)
            )

        latest_date_text = max(
            export_preview["updated_dates"]
        ).replace("-", "")

        original_template_name = st.session_state.get(
            "excel_export_template_name_v070",
            "Manpower.xlsx",
        )
        original_stem = Path(
            original_template_name
        ).stem

        saved_export_mode = st.session_state.get(
            "excel_export_mode_saved_v070",
            "更新現有工程Excel",
        )

        if saved_export_mode == "建立全新工程Excel":
            output_filename = (
                f"{original_stem}_"
                f"{latest_date_text}.xlsx"
            )
            download_label = "下載全新工程Excel"
        else:
            output_filename = (
                f"{original_stem}_updated_"
                f"{latest_date_text}.xlsx"
            )
            download_label = "下載更新後Excel"

        st.download_button(
            download_label,
            data=export_bytes,
            file_name=output_filename,
            mime=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
            type="primary",
            width="stretch",
        )

        st.warning(
            "下載後先在Excel中檢查資料。"
            "更新模式不會覆蓋原始檔；"
            "新工程模式不會包含舊工程歷史資料。"
        )
        